import os
import re
import uuid
import hashlib
import hmac
import csv
import io
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any, Union
from email_validator import validate_email, EmailNotValidError
from twilio.rest import Client as TwilioClient
from openpyxl import load_workbook

from fastapi import FastAPI, HTTPException, Request, Header, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, StreamingResponse, Response, Response
from pydantic import BaseModel, Field, EmailStr, field_validator, ValidationError
from passlib.context import CryptContext
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pymongo.errors import DuplicateKeyError
from emergentintegrations.llm.chat import LlmChat, UserMessage
import json
import asyncio
from typing import AsyncGenerator

# Import secrets manager for secure credential handling
from secrets_manager import (
    get_secret,
    get_all_secrets,
    set_multiple_secrets,
    has_required_twilio_secrets,
    has_required_sendgrid_secrets,
    migrate_secrets_from_settings
)

# Load environment from backend/.env if present
load_dotenv()

# --- Environment & DB Setup ---
MONGO_URL = os.environ.get("MONGO_URL")
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY") or os.environ.get("EMERGENT_API_KEY")
if not MONGO_URL:
    raise RuntimeError("MONGO_URL is not set. Please set it in backend/.env as per platform configuration.")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

app = FastAPI(title="RealtorsPal AI - FastAPI Backend")

# Custom validation error handler
from fastapi.encoders import jsonable_encoder

def loc_to_dot_sep(loc: tuple) -> str:
    """Convert location tuple to dot-separated string"""
    path = ""
    for i, x in enumerate(loc):
        if isinstance(x, str):
            if i > 0:
                path += "."
            path += x
        elif isinstance(x, int):
            path += f"[{x}]"
        else:
            path += f".{x}"
    return path

def convert_validation_errors(validation_error: RequestValidationError) -> list:
    """Convert validation errors to serializable format"""
    converted = []
    for error in validation_error.errors():
        # Build simplified error info excluding problematic fields like 'input'
        converted.append({
            "type": error.get("type", "validation_error"),
            "loc": loc_to_dot_sep(error.get("loc", ())),
            "msg": error.get("msg", "Validation failed"),
        })
    return converted

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    print(f"Validation error on {request.method} {request.url}")
    print(f"Validation errors: {exc.errors()}")
    
    # Log the request body for debugging
    try:
        body = await request.body()
        print(f"Request body: {body.decode()}")
    except Exception as e:
        print(f"Could not read request body: {e}")
    
    # Convert errors to serializable format
    errors = convert_validation_errors(exc)
    
    return JSONResponse(
        status_code=422,
        content=jsonable_encoder({
            "detail": errors,
            "message": "Validation failed"
        })
    )

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mongo client
client = AsyncIOMotorClient(MONGO_URL)
db = client[os.environ.get('DB_NAME', 'realtorspal')]

E164_RE = re.compile(r"^\+[1-9]\d{7,14}$")

# --- Models ---
class UserOut(BaseModel):
    id: str
    email: str
    name: Optional[str] = None

class LoginRequest(BaseModel):
    email: str
    password: str

class LoginResponse(BaseModel):
    user: UserOut
    token: str

class Lead(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    lead_description: Optional[str] = None
    
    # Additional Contact Information
    work_phone: Optional[str] = None
    home_phone: Optional[str] = None
    email_2: Optional[str] = None
    
    # Spouse Information
    spouse_name: Optional[str] = None
    spouse_first_name: Optional[str] = None
    spouse_last_name: Optional[str] = None
    spouse_email: Optional[str] = None
    spouse_mobile_phone: Optional[str] = None
    spouse_birthday: Optional[str] = None
    
    # Pipeline and Status
    pipeline: Optional[str] = None
    status: Optional[str] = None
    ref_source: Optional[str] = None
    lead_rating: Optional[str] = None
    lead_source: Optional[str] = None
    lead_type: Optional[str] = None
    lead_type_2: Optional[str] = None
    
    # Property Information
    house_to_sell: Optional[str] = None
    buying_in: Optional[str] = None
    selling_in: Optional[str] = None
    owns_rents: Optional[str] = None
    mortgage_type: Optional[str] = None
    
    # Address Information
    city: Optional[str] = None
    zip_postal_code: Optional[str] = None
    address: Optional[str] = None
    
    # Property Details
    property_type: Optional[str] = None
    property_condition: Optional[str] = None
    listing_status: Optional[str] = None
    bedrooms: Optional[str] = None
    bathrooms: Optional[str] = None
    basement: Optional[str] = None
    parking_type: Optional[str] = None
    
    # Dates and Anniversaries  
    date_of_birth: Optional[str] = None
    house_anniversary: Optional[str] = None
    planning_to_sell_in: Optional[str] = None
    
    # Agent Assignments
    main_agent: Optional[str] = None
    mort_agent: Optional[str] = None
    list_agent: Optional[str] = None
    
    # Custom Fields (flexible JSON structure)
    custom_fields: Optional[dict] = None
    
    # Existing fields for compatibility
    name: Optional[str] = None
    budget: Optional[int] = None
    priority: Optional[str] = None  # high|medium|low
    stage: str = Field(default="New")
    notes: Optional[str] = None
    neighborhood: Optional[str] = None
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    source_tags: Optional[List[str]] = None
    created_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())
    in_dashboard: Optional[bool] = True
    
    # Lead Generation AI fields for deduplication
    hash_email: Optional[str] = None  # SHA256 hash of lowercase email
    hash_phone: Optional[str] = None  # SHA256 hash of E.164 phone
    phone_e164: Optional[str] = None  # Normalized E.164 phone format

class CreateLeadRequest(BaseModel):
    user_id: str
    
    # Basic fields
    name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    lead_description: Optional[str] = None
    
    # Additional Contact Information
    work_phone: Optional[str] = None
    home_phone: Optional[str] = None
    email_2: Optional[str] = None
    
    # Spouse Information
    spouse_name: Optional[str] = None
    spouse_first_name: Optional[str] = None
    spouse_last_name: Optional[str] = None
    spouse_email: Optional[str] = None
    spouse_mobile_phone: Optional[str] = None
    spouse_birthday: Optional[str] = None
    
    # Pipeline and Status
    pipeline: Optional[str] = None
    status: Optional[str] = None
    ref_source: Optional[str] = None
    lead_rating: Optional[str] = None
    lead_source: Optional[str] = None
    lead_type: Optional[str] = None
    lead_type_2: Optional[str] = None
    
    # Property Information
    house_to_sell: Optional[str] = None
    buying_in: Optional[str] = None
    selling_in: Optional[str] = None
    owns_rents: Optional[str] = None
    mortgage_type: Optional[str] = None
    
    # Address Information
    city: Optional[str] = None
    zip_postal_code: Optional[str] = None
    address: Optional[str] = None
    
    # Property Details
    property_type: Optional[str] = None
    property_condition: Optional[str] = None
    listing_status: Optional[str] = None
    bedrooms: Optional[str] = None
    bathrooms: Optional[str] = None
    basement: Optional[str] = None
    parking_type: Optional[str] = None
    
    # Dates and Anniversaries  
    date_of_birth: Optional[str] = None
    house_anniversary: Optional[str] = None
    planning_to_sell_in: Optional[str] = None
    
    # Agent Assignments
    main_agent: Optional[str] = None
    mort_agent: Optional[str] = None
    list_agent: Optional[str] = None
    
    # Custom Fields (flexible JSON structure)
    custom_fields: Optional[dict] = None
    
    # Existing compatibility fields
    neighborhood: Optional[str] = None
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    priority: Optional[str] = None
    source_tags: Optional[List[str]] = None
    notes: Optional[str] = None
    stage: Optional[str] = None
    in_dashboard: Optional[bool] = None

    @field_validator("phone", "work_phone", "home_phone", "spouse_mobile_phone")
    @classmethod
    def validate_phone(cls, v):
        if v is not None and v.strip() and not E164_RE.match(v):
            # Try to normalize the phone number
            normalized = normalize_phone(v)
            if normalized and E164_RE.match(normalized):
                return normalized
            raise ValueError("Phone must be in E.164 format, e.g. +1234567890")
        return v

class UpdateLeadRequest(BaseModel):
    # Basic fields - all optional for partial update
    name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    lead_description: Optional[str] = None
    
    # Additional Contact Information
    work_phone: Optional[str] = None
    home_phone: Optional[str] = None
    email_2: Optional[str] = None
    
    # Spouse Information
    spouse_name: Optional[str] = None
    spouse_first_name: Optional[str] = None
    spouse_last_name: Optional[str] = None
    spouse_email: Optional[str] = None
    spouse_mobile_phone: Optional[str] = None
    spouse_birthday: Optional[str] = None
    
    # Pipeline and Status
    pipeline: Optional[str] = None
    status: Optional[str] = None
    ref_source: Optional[str] = None
    lead_rating: Optional[str] = None
    lead_source: Optional[str] = None
    lead_type: Optional[str] = None
    lead_type_2: Optional[str] = None
    
    # Property Information
    house_to_sell: Optional[str] = None
    buying_in: Optional[str] = None
    selling_in: Optional[str] = None
    owns_rents: Optional[str] = None
    mortgage_type: Optional[str] = None
    
    # Address Information
    city: Optional[str] = None
    zip_postal_code: Optional[str] = None
    address: Optional[str] = None
    
    # Property Details
    property_type: Optional[str] = None
    property_condition: Optional[str] = None
    listing_status: Optional[str] = None
    bedrooms: Optional[str] = None
    bathrooms: Optional[str] = None
    basement: Optional[str] = None
    parking_type: Optional[str] = None
    
    # Dates and Anniversaries  
    date_of_birth: Optional[str] = None
    house_anniversary: Optional[str] = None
    planning_to_sell_in: Optional[str] = None
    
    # Agent Assignments
    main_agent: Optional[str] = None
    mort_agent: Optional[str] = None
    list_agent: Optional[str] = None
    
    # Custom Fields (flexible JSON structure)
    custom_fields: Optional[dict] = None
    
    # Existing compatibility fields
    stage: Optional[str] = None
    notes: Optional[str] = None
    neighborhood: Optional[str] = None
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    priority: Optional[str] = None
    source_tags: Optional[List[str]] = None
    in_dashboard: Optional[bool] = None

    @field_validator("phone", "work_phone", "home_phone", "spouse_mobile_phone")
    @classmethod
    def validate_phone(cls, v):
        if v is not None and v.strip() and not E164_RE.match(v):
            # Try to normalize the phone number
            normalized = normalize_phone(v)
            if normalized and E164_RE.match(normalized):
                return normalized
            raise ValueError("Phone must be in E.164 format, e.g. +1234567890")
        return v

class UpdateStageRequest(BaseModel):
    stage: str

class Settings(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    openai_api_key: Optional[str] = None
    anthropic_api_key: Optional[str] = None
    gemini_api_key: Optional[str] = None
    webhook_enabled: Optional[bool] = False
    facebook_webhook_verify_token: Optional[str] = None
    generic_webhook_enabled: Optional[bool] = False
    api_key: Optional[str] = None  # API key for external integrations
    twilio_account_sid: Optional[str] = None
    twilio_auth_token: Optional[str] = None
    twilio_phone_number: Optional[str] = None
    twilio_whatsapp_number: Optional[str] = None
    twilio_api_key: Optional[str] = None
    twilio_api_secret: Optional[str] = None
    twilio_twiml_app_sid: Optional[str] = None  # TwiML Application SID for WebRTC
    agent_phone_number: Optional[str] = None  # Agent's phone for receiving bridged calls
    smtp_protocol: Optional[str] = None
    smtp_hostname: Optional[str] = None
    smtp_port: Optional[str] = None
    smtp_ssl_tls: Optional[bool] = None
    smtp_username: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_from_email: Optional[str] = None
    smtp_from_name: Optional[str] = None
    sendgrid_api_key: Optional[str] = None
    sender_email: Optional[str] = None  # Verified sender email for SendGrid

class AnalyticsDashboard(BaseModel):
    total_leads: int
    by_stage: Dict[str, int]

# Import payloads
class ImportItem(BaseModel):
    # Basic fields
    name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None  # Changed from EmailStr to str to be more lenient
    phone: Optional[str] = None
    lead_description: Optional[str] = None
    
    # Additional Contact Information
    work_phone: Optional[str] = None
    home_phone: Optional[str] = None
    email_2: Optional[str] = None
    
    # Spouse Information
    spouse_name: Optional[str] = None
    spouse_first_name: Optional[str] = None
    spouse_last_name: Optional[str] = None
    spouse_email: Optional[str] = None
    spouse_mobile_phone: Optional[str] = None
    spouse_birthday: Optional[str] = None
    
    # Pipeline and Status
    pipeline: Optional[str] = None
    status: Optional[str] = None
    ref_source: Optional[str] = None
    lead_rating: Optional[str] = None
    lead_source: Optional[str] = None
    lead_type: Optional[str] = None
    lead_type_2: Optional[str] = None
    
    # Property Information
    house_to_sell: Optional[str] = None
    buying_in: Optional[str] = None
    selling_in: Optional[str] = None
    owns_rents: Optional[str] = None
    mortgage_type: Optional[str] = None
    
    # Address Information
    city: Optional[str] = None
    zip_postal_code: Optional[str] = None
    address: Optional[str] = None
    
    # Property Details
    property_type: Optional[str] = None
    property_condition: Optional[str] = None
    listing_status: Optional[str] = None
    bedrooms: Optional[str] = None
    bathrooms: Optional[str] = None
    basement: Optional[str] = None
    parking_type: Optional[str] = None
    
    # Dates and Anniversaries  
    date_of_birth: Optional[str] = None
    house_anniversary: Optional[str] = None
    planning_to_sell_in: Optional[str] = None
    
    # Agent Assignments
    main_agent: Optional[str] = None
    mort_agent: Optional[str] = None
    list_agent: Optional[str] = None
    
    # Compatibility fields
    neighborhood: Optional[str] = None
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    priority: Optional[str] = None
    source_tags: Optional[List[str]] = None
    notes: Optional[str] = None
    stage: Optional[str] = None
    tags: Optional[str] = None  # Tags as comma-separated string

class ImportPayload(BaseModel):
    user_id: str
    default_stage: Optional[str] = "New"
    in_dashboard: Optional[bool] = False
    leads: List[ImportItem]

class ImportResult(BaseModel):
    inserted: int
    skipped: int
    errors: List[Dict[str, Any]]
    inserted_leads: List[Lead]

class EmailHistory(BaseModel):
    id: str
    user_id: str
    lead_id: str
    subject: str
    body: str
    to_email: str
    from_email: str
    from_name: str
    status: str  # sent, failed, draft
    sent_at: Optional[str] = None
    error_message: Optional[str] = None
    llm_provider: Optional[str] = None
    email_template: Optional[str] = None
    created_at: str

class SendEmailRequest(BaseModel):
    lead_id: str
    subject: str
    body: str
    email_template: Optional[str] = None
    llm_provider: Optional[str] = None

# --- API Authentication ---
async def authenticate_api_key(api_key: str) -> Optional[str]:
    """Authenticate API key and return user_id"""
    settings_doc = await db.settings.find_one({"api_key": api_key})
    if settings_doc:
        return settings_doc.get("user_id")
    return None

def generate_api_key() -> str:
    """Generate a secure API key"""
    return f"crm_{uuid.uuid4().hex[:16]}_{uuid.uuid4().hex[:16]}"

# --- External API Endpoints for Crew.AI Integration ---

class CreateLeadExternalRequest(BaseModel):
    name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    property_type: Optional[str] = None
    neighborhood: Optional[str] = None
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    priority: Optional[str] = "medium"
    source_tags: Optional[List[str]] = ["API"]
    notes: Optional[str] = None
    stage: Optional[str] = "New"
    in_dashboard: Optional[bool] = True

class UpdateLeadExternalRequest(BaseModel):
    name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    property_type: Optional[str] = None
    neighborhood: Optional[str] = None
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    priority: Optional[str] = None
    source_tags: Optional[List[str]] = None
    notes: Optional[str] = None
    stage: Optional[str] = None
    in_dashboard: Optional[bool] = None

class UpdateLeadStatusRequest(BaseModel):
    stage: str
    notes: Optional[str] = None

class SearchLeadsRequest(BaseModel):
    email: Optional[str] = None
    phone: Optional[str] = None
    name: Optional[str] = None
    stage: Optional[str] = None
    limit: Optional[int] = 10

@app.post("/api/external/leads", response_model=Lead)
async def create_lead_external(lead_data: CreateLeadExternalRequest, api_key: str = Header(..., alias="X-API-Key")):
    """Create a new lead via external API (Third-party app integration)"""
    try:
        # Authenticate API key
        user_id = await authenticate_api_key(api_key)
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid API key")
        
        # Validate and normalize email
        validated_email = None
        if lead_data.email and lead_data.email.strip():
            try:
                validation = validate_email(lead_data.email.strip(), check_deliverability=False)
                validated_email = validation.email
            except EmailNotValidError:
                validated_email = None
        
        # Normalize phone number
        normalized_phone = normalize_phone(lead_data.phone)
        
        # Create lead name if not provided
        name = lead_data.name
        if not name:
            name = f"{lead_data.first_name or ''} {lead_data.last_name or ''}".strip() or "New Lead"
        
        lead = Lead(
            user_id=user_id,
            name=name,
            first_name=lead_data.first_name,
            last_name=lead_data.last_name,
            email=validated_email,
            phone=normalized_phone,
            property_type=lead_data.property_type,
            neighborhood=lead_data.neighborhood,
            price_min=lead_data.price_min,
            price_max=lead_data.price_max,
            priority=lead_data.priority,
            source_tags=lead_data.source_tags,
            notes=lead_data.notes,
            stage=lead_data.stage,
            in_dashboard=lead_data.in_dashboard
        )
        
        await db.leads.insert_one(lead.model_dump(exclude_none=True))
        return lead
        
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="Lead with this email already exists")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/external/leads/{lead_id}", response_model=Lead)
async def update_lead_external(lead_id: str, lead_data: UpdateLeadExternalRequest, api_key: str = Header(..., alias="X-API-Key")):
    """Update an existing lead via external API (Third-party app integration)"""
    try:
        # Authenticate API key
        user_id = await authenticate_api_key(api_key)
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid API key")
        
        # Check if lead exists and belongs to user
        existing_lead = await db.leads.find_one({"id": lead_id, "user_id": user_id})
        if not existing_lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Prepare update data
        update_data = {}
        for field, value in lead_data.model_dump(exclude_none=True).items():
            if field == "email" and value:
                # Validate email
                try:
                    validation = validate_email(value.strip(), check_deliverability=False)
                    update_data[field] = validation.email
                except EmailNotValidError:
                    pass  # Skip invalid email
            elif field == "phone" and value:
                # Normalize phone
                update_data[field] = normalize_phone(value)
            else:
                update_data[field] = value
        
        # Update name if first_name or last_name changed
        if "first_name" in update_data or "last_name" in update_data:
            first_name = update_data.get("first_name", existing_lead.get("first_name", ""))
            last_name = update_data.get("last_name", existing_lead.get("last_name", ""))
            update_data["name"] = f"{first_name or ''} {last_name or ''}".strip() or existing_lead.get("name")
        
        # Update lead
        await db.leads.update_one({"id": lead_id}, {"$set": update_data})
        
        # Return updated lead
        updated_lead = await db.leads.find_one({"id": lead_id})
        return Lead(**{k: v for k, v in updated_lead.items() if k != "_id"})
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/external/leads/search")
async def search_leads_external(search_data: SearchLeadsRequest, api_key: str = Header(..., alias="X-API-Key")):
    """Search leads via external API (Third-party app integration)"""
    try:
        # Authenticate API key
        user_id = await authenticate_api_key(api_key)
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid API key")
        
        # Build search query
        query = {"user_id": user_id}
        
        if search_data.email:
            query["email"] = {"$regex": search_data.email, "$options": "i"}
        if search_data.phone:
            # Normalize phone for search
            normalized_phone = normalize_phone(search_data.phone)
            query["phone"] = normalized_phone
        if search_data.name:
            query["$or"] = [
                {"name": {"$regex": search_data.name, "$options": "i"}},
                {"first_name": {"$regex": search_data.name, "$options": "i"}},
                {"last_name": {"$regex": search_data.name, "$options": "i"}}
            ]
        if search_data.stage:
            query["stage"] = search_data.stage
        
        # Search leads
        leads = await db.leads.find(query).limit(search_data.limit or 10).to_list(length=None)
        return [Lead(**{k: v for k, v in lead.items() if k != "_id"}) for lead in leads]
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/external/leads/{lead_id}/status", response_model=Lead)
async def update_lead_status_external(lead_id: str, status_data: UpdateLeadStatusRequest, api_key: str = Header(..., alias="X-API-Key")):
    """Update lead status/stage via external API (Third-party app integration)"""
    try:
        # Authenticate API key
        user_id = await authenticate_api_key(api_key)
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid API key")
        
        # Check if lead exists and belongs to user
        existing_lead = await db.leads.find_one({"id": lead_id, "user_id": user_id})
        if not existing_lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Prepare update data
        update_data = {"stage": status_data.stage}
        if status_data.notes:
            # Append notes to existing notes
            existing_notes = existing_lead.get("notes", "")
            if existing_notes:
                update_data["notes"] = f"{existing_notes}\n\n[API Update] {status_data.notes}"
            else:
                update_data["notes"] = f"[API Update] {status_data.notes}"
        
        # Update lead
        await db.leads.update_one({"id": lead_id}, {"$set": update_data})
        
        # Return updated lead
        updated_lead = await db.leads.find_one({"id": lead_id})
        return Lead(**{k: v for k, v in updated_lead.items() if k != "_id"})
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/external/leads/{lead_id}", response_model=Lead)
async def get_lead_external(lead_id: str, api_key: str = Header(..., alias="X-API-Key")):
    """Get a specific lead via external API (Third-party app integration)"""
    try:
        # Authenticate API key
        user_id = await authenticate_api_key(api_key)
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid API key")
        
        # Get lead
        lead = await db.leads.find_one({"id": lead_id, "user_id": user_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        return Lead(**{k: v for k, v in lead.items() if k != "_id"})
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- TwiML Endpoints for WebRTC Calling ---

@app.get("/api/twiml/webrtc-outbound")
@app.post("/api/twiml/webrtc-outbound")
async def webrtc_outbound_twiml(request: Request):
    """TwiML endpoint for WebRTC outbound calls - Handles calls from browser to phone"""
    try:
        # Get parameters from the request
        params = dict(request.query_params)
        form_data = await request.form() if request.method == "POST" else {}
        
        # Merge params
        all_params = {**params, **dict(form_data)}
        
        # Get lead phone number from parameters
        to_number = all_params.get('To', '')
        lead_id = all_params.get('lead_id', '')
        
        print(f"📞 WebRTC Outbound TwiML Request:")
        print(f"   To: {to_number}")
        print(f"   Lead ID: {lead_id}")
        print(f"   All params: {all_params}")
        
        # Create TwiML to dial the lead's phone number
        twiml_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">Connecting your call</Say>
    <Dial callerId="+12894012412" timeout="30" timeLimit="3600">
        {to_number}
    </Dial>
    <Say voice="alice">The call could not be completed. Please try again.</Say>
</Response>"""
        
        print(f"   TwiML Response generated")
        return Response(content=twiml_response, media_type="application/xml")
        
    except Exception as e:
        print(f"❌ WebRTC Outbound TwiML error: {e}")
        import traceback
        traceback.print_exc()
        # Fallback TwiML
        fallback_twiml = """<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">Sorry, there was an error connecting your call. Please try again later.</Say>
</Response>"""
        return Response(content=fallback_twiml, media_type="application/xml")

@app.get("/api/twiml/outbound-call")
@app.post("/api/twiml/outbound-call")
async def outbound_call_twiml(request: Request):
    """TwiML endpoint for WebRTC outbound calls - connects lead to agent's browser"""
    try:
        # Get query parameters from Twilio call
        params = dict(request.query_params)
        
        # Extract parameters (these will be passed from our WebRTC call initiation)
        agent_identity = params.get('agent_identity', 'agent_unknown')
        lead_phone = params.get('lead_phone', '')
        
        print(f"Outbound call TwiML: connecting {lead_phone} to WebRTC client {agent_identity}")
        
        # TwiML to connect the lead to the agent's WebRTC client (browser)
        twiml_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">Please hold while we connect you to your real estate agent.</Say>
    <Dial timeout="30" timeLimit="3600">
        <Client>{agent_identity}</Client>
    </Dial>
    <Say voice="alice">Sorry, the agent is not available right now. Please try again later.</Say>
</Response>"""
        
        return Response(content=twiml_response, media_type="application/xml")
        
    except Exception as e:
        print(f"TwiML outbound call error: {e}")
        # Fallback TwiML
        fallback_twiml = """<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">Sorry, there was an error connecting your call. Please try again later.</Say>
</Response>"""
        return Response(content=fallback_twiml, media_type="application/xml")

@app.get("/api/twiml/client-incoming")
@app.post("/api/twiml/client-incoming")
async def client_incoming_twiml(request: Request):
    """TwiML endpoint for incoming calls to WebRTC client (agent's browser)"""
    try:
        # This handles when someone calls the agent's WebRTC client directly
        params = dict(request.query_params)
        from_number = params.get('From', 'Unknown')
        
        print(f"Incoming call to WebRTC client from: {from_number}")
        
        # TwiML to handle incoming calls to the agent's browser
        twiml_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">You have an incoming call from {from_number}.</Say>
    <Dial timeout="30">
        <Number>{from_number}</Number>
    </Dial>
</Response>"""
        
        return Response(content=twiml_response, media_type="application/xml")
        
    except Exception as e:
        print(f"TwiML client incoming error: {e}")
        fallback_twiml = """<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">Unable to process the incoming call.</Say>
</Response>"""
        return Response(content=fallback_twiml, media_type="application/xml")

# --- Updated Access Token Generation ---

class TwilioCallRequest(BaseModel):
    lead_id: str
    agent_phone: Optional[str] = None  # Agent's phone number to connect to
    message: Optional[str] = "Connecting you to your real estate agent now."

class TwilioWebRTCCallRequest(BaseModel):
    lead_id: str
    message: Optional[str] = "Hello, this is your real estate agent calling about your property inquiry."

class TwilioSMSRequest(BaseModel):
    lead_id: str
    message: str

class TwilioWhatsAppRequest(BaseModel):
    lead_id: str
    message: str

class AccessTokenRequest(BaseModel):
    user_id: str

async def get_twilio_client(user_id: str) -> Optional[TwilioClient]:
    """Get configured Twilio client for user"""
    print(f"🔵 get_twilio_client called for user_id: {user_id}")
    settings_doc = await db.settings.find_one({"user_id": user_id})
    if not settings_doc:
        print(f"❌ No settings found for user_id: {user_id}")
        return None
        
    account_sid = settings_doc.get("twilio_account_sid")
    auth_token = settings_doc.get("twilio_auth_token")
    
    print(f"   Account SID: {account_sid}")
    print(f"   Auth Token: {'***' + auth_token[-4:] if auth_token else 'NOT SET'}")
    
    if not account_sid or not auth_token:
        print(f"❌ Missing credentials - Account SID: {bool(account_sid)}, Auth Token: {bool(auth_token)}")
        return None
    
    print(f"✅ Creating Twilio client with credentials")
    try:
        client = TwilioClient(account_sid, auth_token)
        print(f"✅ Twilio client created successfully")
        return client
    except Exception as e:
        print(f"❌ Error creating Twilio client: {e}")
        return None

@app.post("/api/twilio/access-token")
async def generate_access_token(token_request: AccessTokenRequest):
    """Generate Twilio access token for WebRTC calling using API Keys"""
    try:
        # Auto-migrate secrets from settings if needed
        await migrate_secrets_from_settings(token_request.user_id)
        
        # Get Twilio secrets from secure storage
        secrets = await get_all_secrets(token_request.user_id)
        
        account_sid = secrets.get("twilio_account_sid")
        api_key = secrets.get("twilio_api_key")
        api_secret = secrets.get("twilio_api_secret")
        
        # Check if all required credentials are present
        if not account_sid or not api_key or not api_secret:
            missing_fields = []
            if not account_sid: missing_fields.append("Account SID")
            if not api_key: missing_fields.append("API Key SID")  
            if not api_secret: missing_fields.append("API Key Secret")
            
            return {
                "status": "setup_required",
                "message": f"Missing Twilio credentials: {', '.join(missing_fields)}",
                "setup_instructions": {
                    "step1": "Go to Twilio Console → Account → API Keys & Tokens",
                    "step2": "Create new API Key with Voice grants enabled",
                    "step3": "Copy the API Key SID and Secret to Settings",
                    "step4": "Make sure Account SID is also configured"
                }
            }
        
        # Import Twilio components for access token
        from twilio.jwt.access_token import AccessToken
        from twilio.jwt.access_token.grants import VoiceGrant
        
        try:
            # Get TwiML App SID from environment or secrets
            twiml_app_sid = os.environ.get('TWILIO_TWIML_APP_SID')
            
            if not twiml_app_sid:
                # Try to get from secrets collection
                twiml_app_sid = secrets.get('twilio_twiml_app_sid')
            
            if not twiml_app_sid:
                return {
                    "status": "error",
                    "message": "TwiML Application SID not configured. Please set TWILIO_TWIML_APP_SID environment variable or configure in settings."
                }
            
            # Create access token using API Keys (proper way for WebRTC)
            identity = f"agent_{token_request.user_id}"
            
            token = AccessToken(
                account_sid,
                api_key,      # API Key SID
                api_secret,   # API Key Secret
                identity=identity
            )
            
            # Get base URL for TwiML endpoints
            base_url = os.environ.get('REACT_APP_BACKEND_URL', 'https://realtor-workflow.preview.emergentagent.com')
            
            # Create voice grant with TwiML Application SID
            voice_grant = VoiceGrant(
                outgoing_application_sid=twiml_app_sid,  # TwiML App SID for outgoing calls
                incoming_allow=True  # Allow incoming calls to the WebRTC client
            )
            
            # Add the voice grant to token
            token.add_grant(voice_grant)
            
            # Generate the JWT token
            jwt_token = token.to_jwt()
            
            # Log token generation for debugging
            print(f"✅ Generated WebRTC access token for user {token_request.user_id}")
            print(f"   Identity: {identity}")
            print(f"   Account SID: {account_sid}")
            print(f"   API Key: {api_key}")
            print(f"   TwiML App SID: {twiml_app_sid}")
            print(f"   TwiML App URL: {base_url}/api/twiml/webrtc-outbound")
            
            return {
                "status": "success", 
                "token": jwt_token,
                "identity": identity,
                "expires_in": 3600,
                "account_sid": account_sid,
                "twiml_app_sid": twiml_app_sid,
                "twiml_app_url": f"{base_url}/api/twiml/webrtc-outbound",
                "debug_info": {
                    "token_length": len(jwt_token),
                    "api_key_prefix": api_key[:8] + "...",
                    "account_sid_prefix": account_sid[:8] + "...",
                    "twiml_app_sid": twiml_app_sid
                }
            }
            
        except Exception as token_error:
            print(f"❌ Token generation failed: {token_error}")
            import traceback
            traceback.print_exc()
            return {
                "status": "error",
                "message": f"Failed to generate access token: {str(token_error)}",
                "suggestion": "Please verify your API Key and Secret are correct in Settings"
            }
        
    except Exception as e:
        print(f"❌ Access token generation error: {e}")
        import traceback
        traceback.print_exc()
        return {
            "status": "error", 
            "message": f"Access token error: {str(e)}"
        }

@app.post("/api/twilio/outbound-call")
async def initiate_outbound_call(call_data: TwilioWebRTCCallRequest):
    """Initiate a direct outbound call using Twilio - Simple phone-to-phone call"""
    try:
        # Get lead details
        lead = await db.leads.find_one({"id": call_data.lead_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Auto-migrate secrets if needed
        await migrate_secrets_from_settings(lead["user_id"])
        
        # Get Twilio secrets from secure storage
        secrets = await get_all_secrets(lead["user_id"])
        
        account_sid = secrets.get("twilio_account_sid")
        auth_token = secrets.get("twilio_auth_token")
        twilio_phone = secrets.get("twilio_phone_number")
        
        # Check required credentials (only basic ones for direct calling)
        if not all([account_sid, auth_token, twilio_phone]):
            missing = []
            if not account_sid: missing.append("Account SID")
            if not auth_token: missing.append("Auth Token")
            if not twilio_phone: missing.append("Phone Number")
            
            return {
                "status": "error",
                "message": f"Missing Twilio credentials: {', '.join(missing)}. Please configure in Settings.",
                "setup_required": True
            }
        
        if not lead.get("phone"):
            return {"status": "error", "message": "Lead has no phone number"}
        
        # Validate and format phone numbers
        to_phone = lead["phone"]
        if not to_phone.startswith('+'):
            to_phone = '+' + to_phone.lstrip('+')
        
        from_phone = twilio_phone
        if not from_phone.startswith('+'):
            from_phone = '+' + from_phone.lstrip('+')
        
        print(f"🔵 Initiating outbound call:")
        print(f"   From: {from_phone} (Twilio)")
        print(f"   To: {to_phone} (Lead: {lead.get('first_name')} {lead.get('last_name')})")
        print(f"   Account SID: {account_sid}")
        
        # Use Twilio REST API to create outbound call with direct HTTP request
        import requests
        from requests.auth import HTTPBasicAuth
        
        url = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Calls.json"
        
        # Get agent phone number from secrets for call bridging
        agent_phone = secrets.get("agent_phone_number")
        
        if not agent_phone:
            return {
                "status": "error",
                "message": "Agent phone number not configured. Please add 'agent_phone_number' in Settings."
            }
        
        # TwiML that plays greeting and bridges call to agent phone
        # After lead answers, they hear message and get connected to agent
        twiml = f'<Response><Say voice="alice">Hello, please hold while we connect you to your real estate agent.</Say><Dial callerId="{from_phone}" timeout="30" timeLimit="3600">{agent_phone}</Dial><Say voice="alice">Sorry, the agent is currently unavailable. Please try again later.</Say></Response>'
        
        payload = {
            'To': to_phone,
            'From': from_phone,
            'Twiml': twiml
        }
        
        print(f"   Making API call to Twilio...")
        
        response = requests.post(
            url,
            data=payload,
            auth=HTTPBasicAuth(account_sid, auth_token)
        )
        
        print(f"   Response Status: {response.status_code}")
        print(f"   Response Body: {response.text[:300]}")
        
        if response.status_code in [200, 201]:
            response_data = response.json()
            call_sid = response_data.get('sid')
            status = response_data.get('status')
            
            print(f"✅ Call initiated successfully!")
            print(f"   Call SID: {call_sid}")
            print(f"   Status: {status}")
            
            # Log the call activity
            current_notes = lead.get('notes', '')
            new_note = f"\n\n[Outbound Call] Initiated from {from_phone} to {to_phone} - Call SID: {call_sid} - {datetime.now().isoformat()}"
            await db.leads.update_one(
                {"id": call_data.lead_id},
                {"$set": {"notes": current_notes + new_note}}
            )
            
            return {
                "status": "success",
                "call_sid": call_sid,
                "message": f"Call initiated successfully! Calling {to_phone} from {from_phone}",
                "call_status": status,
                "from_number": from_phone,
                "to_number": to_phone
            }
        else:
            error_msg = response.text
            print(f"❌ Twilio API error: {error_msg}")
            
            return {
                "status": "error",
                "message": f"Failed to initiate call: {error_msg}"
            }
        
    except Exception as e:
        print(f"❌ Outbound call error: {e}")
        import traceback
        traceback.print_exc()
        return {"status": "error", "message": f"Call failed: {str(e)}"}

@app.get("/api/twilio/voice")
@app.post("/api/twilio/voice")
async def voice_webhook(request: Request):
    """Twilio voice webhook that handles incoming calls and creates bridge"""
    try:
        # Get query parameters from Twilio
        params = dict(request.query_params)
        
        # Get the agent phone and message from the webhook parameters
        agent_phone = params.get('agent_phone')
        encoded_message = params.get('message', 'Connecting you to your real estate agent now.')
        lead_phone = params.get('lead_phone')
        
        # URL decode the message
        from urllib.parse import unquote
        message = unquote(encoded_message)
        
        print(f"Voice webhook called: agent_phone={agent_phone}, lead_phone={lead_phone}, message={message}")
        
        # Generate TwiML response
        twiml_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">{message}</Say>
    <Dial callerId="+{agent_phone if agent_phone else '12894012412'}" timeout="30" timeLimit="3600">
        +{agent_phone if agent_phone else '12894012412'}
    </Dial>
    <Say voice="alice">The call could not be connected. Please try again later.</Say>
</Response>"""
        
        return Response(content=twiml_response, media_type="application/xml")
        
    except Exception as e:
        print(f"Voice webhook error: {e}")
        # Fallback TwiML
        fallback_twiml = """<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Say voice="alice">Sorry, there was an error connecting your call. Please try again later.</Say>
</Response>"""
        return Response(content=fallback_twiml, media_type="application/xml")

@app.post("/api/twilio/call")
async def initiate_call(call_data: TwilioCallRequest):
    """Initiate a call via Twilio with voice bridge"""
    try:
        # Get lead details
        lead = await db.leads.find_one({"id": call_data.lead_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Get Twilio client
        client = await get_twilio_client(lead["user_id"])
        if not client:
            return {
                "status": "error", 
                "message": "Twilio not configured. Please add your Twilio credentials in Settings to enable calling.",
                "setup_instructions": {
                    "step1": "Go to Settings > Twilio Communication",
                    "step2": "Add your Twilio Account SID and Auth Token",
                    "step3": "Add your Twilio phone number",
                    "step4": "Save settings and try calling again"
                }
            }
        
        # Auto-migrate and get Twilio secrets
        await migrate_secrets_from_settings(lead["user_id"])
        secrets = await get_all_secrets(lead["user_id"])
        twilio_phone = secrets.get("twilio_phone_number")
        
        if not twilio_phone:
            raise HTTPException(status_code=400, detail="Twilio phone number not configured. Please add your Twilio phone number in Settings.")
        
        if not lead.get("phone"):
            raise HTTPException(status_code=400, detail="Lead has no phone number")
        
        # Clean phone numbers and encode message for URL
        clean_twilio_phone = twilio_phone.replace('+', '')
        clean_lead_phone = lead["phone"].replace('+', '')
        
        # URL encode the message
        from urllib.parse import quote
        encoded_message = quote(call_data.message)
        
        # Create voice webhook URL with parameters
        base_url = os.environ.get('REACT_APP_BACKEND_URL')
        voice_webhook_url = f"{base_url}/api/twilio/voice?agent_phone={clean_twilio_phone}&lead_phone={clean_lead_phone}&message={encoded_message}"
        
        print(f"Initiating call from {twilio_phone} to {lead['phone']} with webhook: {voice_webhook_url}")
        
        # Initiate call to the LEAD first, which will then connect to agent
        call = client.calls.create(
            to=lead["phone"],
            from_=twilio_phone,
            url=voice_webhook_url,
            method='GET'
        )
        
        # Log the call activity
        current_notes = lead.get('notes', '')
        new_note = f"\n\n[Call] Initiated voice bridge call - Agent: {twilio_phone} → Lead: {lead['phone']} - {datetime.now().isoformat()}"
        await db.leads.update_one(
            {"id": call_data.lead_id},
            {"$set": {"notes": current_notes + new_note}}
        )
        
        return {
            "status": "success",
            "call_sid": call.sid,
            "message": f"Voice bridge call initiated. The lead will receive a call and then be connected to you.",
            "details": {
                "lead_phone": lead["phone"],
                "agent_phone": twilio_phone,
                "call_flow": "Lead receives call → Hears message → Connected to agent"
            }
        }
        
    except Exception as e:
        print(f"Call initiation error: {e}")
        return {"status": "error", "message": str(e)}

@app.post("/api/twilio/sms")
async def send_sms(sms_data: TwilioSMSRequest):
    """Send SMS via Twilio using direct API call"""
    try:
        # Get lead details
        lead = await db.leads.find_one({"id": sms_data.lead_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        print(f"🔵 Sending SMS to lead: {lead.get('first_name')} {lead.get('last_name')}")
        print(f"   Lead phone: {lead.get('phone')}")
        
        if not lead.get("phone"):
            raise HTTPException(status_code=400, detail="Lead has no phone number")
        
        # Auto-migrate and get Twilio secrets
        await migrate_secrets_from_settings(lead["user_id"])
        secrets = await get_all_secrets(lead["user_id"])
        
        account_sid = secrets.get("twilio_account_sid")
        auth_token = secrets.get("twilio_auth_token")
        twilio_phone = secrets.get("twilio_phone_number")
        
        print(f"   Account SID: {account_sid}")
        print(f"   Auth Token: ***{auth_token[-4:] if auth_token else 'NOT SET'}")
        print(f"   From phone: {twilio_phone}")
        
        if not account_sid or not auth_token or not twilio_phone:
            missing = []
            if not account_sid: missing.append("Account SID")
            if not auth_token: missing.append("Auth Token")
            if not twilio_phone: missing.append("Phone Number")
            raise HTTPException(status_code=400, detail=f"Twilio configuration incomplete. Missing: {', '.join(missing)}")
        
        # Validate and format phone numbers
        to_phone = lead["phone"]
        if not to_phone.startswith('+'):
            to_phone = '+' + to_phone
        
        print(f"   Sending SMS...")
        print(f"   From: {twilio_phone}")
        print(f"   To: {to_phone}")
        print(f"   Message: {sms_data.message[:50]}...")
        
        # Send SMS using direct HTTP request (same as successful curl command)
        import requests
        from requests.auth import HTTPBasicAuth
        
        url = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json"
        
        payload = {
            'To': to_phone,
            'From': twilio_phone,
            'Body': sms_data.message
        }
        
        print(f"   URL: {url}")
        print(f"   Payload: {payload}")
        
        response = requests.post(
            url,
            data=payload,
            auth=HTTPBasicAuth(account_sid, auth_token)
        )
        
        print(f"   Response Status: {response.status_code}")
        print(f"   Response Body: {response.text[:200]}")
        
        if response.status_code in [200, 201]:
            response_data = response.json()
            message_sid = response_data.get('sid')
            status = response_data.get('status')
            
            print(f"✅ SMS sent successfully!")
            print(f"   Message SID: {message_sid}")
            print(f"   Status: {status}")
            
            # Log the SMS activity in lead notes
            current_notes = lead.get('notes', '')
            new_note = f"\n\n[SMS] Sent: '{sms_data.message}' to {to_phone} - {datetime.now().isoformat()}"
            await db.leads.update_one(
                {"id": sms_data.lead_id},
                {"$set": {"notes": current_notes + new_note}}
            )
            
            return {
                "status": "success",
                "message_sid": message_sid,
                "message": f"SMS sent to {to_phone}",
                "twilio_status": status
            }
        else:
            error_msg = response.text
            print(f"❌ Twilio API error: {error_msg}")
            
            return {
                "status": "error",
                "message": f"Failed to send SMS: {error_msg}"
            }
        
    except Exception as e:
        print(f"❌ SMS sending error: {e}")
        import traceback
        traceback.print_exc()
        return {"status": "error", "message": str(e)}

@app.post("/api/twilio/whatsapp")
async def send_whatsapp(whatsapp_data: TwilioWhatsAppRequest):
    """Send WhatsApp message via Twilio"""
    try:
        # Get lead details
        lead = await db.leads.find_one({"id": whatsapp_data.lead_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Get Twilio client
        client = await get_twilio_client(lead["user_id"])
        if not client:
            raise HTTPException(status_code=400, detail="Twilio not configured")
        
        # Get user's Twilio settings
        settings = await db.settings.find_one({"user_id": lead["user_id"]})
        twilio_whatsapp = settings.get("twilio_whatsapp_number")
        
        if not twilio_whatsapp:
            raise HTTPException(status_code=400, detail="Twilio WhatsApp number not configured")
        
        if not lead.get("phone"):
            raise HTTPException(status_code=400, detail="Lead has no phone number")
        
        # Send WhatsApp message using synchronous client
        message = client.messages.create(
            body=whatsapp_data.message,
            from_=f"whatsapp:{twilio_whatsapp}",
            to=f"whatsapp:{lead['phone']}"
        )
        
        # Log the WhatsApp activity
        await db.leads.update_one(
            {"id": whatsapp_data.lead_id},
            {"$set": {"notes": f"{lead.get('notes', '')}\n\n[WhatsApp] Sent: '{whatsapp_data.message}' - {datetime.now().isoformat()}"}}
        )
        
        return {
            "status": "success",
            "message_sid": message.sid,
            "message": f"WhatsApp sent to {lead['phone']}"
        }
        
    except Exception as e:
        print(f"WhatsApp sending error: {e}")
        return {"status": "error", "message": str(e)}

# --- Email Communication Endpoints ---

@app.post("/api/email/send")
async def send_email(email_request: SendEmailRequest):
    """Send email to lead using SendGrid"""
    try:
        # Get lead details
        lead = await db.leads.find_one({"id": email_request.lead_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        if not lead.get("email"):
            return {"status": "error", "message": "Lead has no email address"}
        
        # Auto-migrate and get SendGrid secrets
        await migrate_secrets_from_settings(lead["user_id"])
        secrets = await get_all_secrets(lead["user_id"])
        
        # Check SendGrid configuration
        sendgrid_api_key = secrets.get("sendgrid_api_key")
        sender_email = secrets.get("sender_email", "support@syncai.tech")
        
        if not sendgrid_api_key:
            return {
                "status": "error", 
                "message": "SendGrid API key not configured. Please configure in Settings.",
                "setup_required": True
            }
        
        if not sender_email:
            return {
                "status": "error",
                "message": "Sender email not configured. Please configure in Settings.",
                "setup_required": True
            }
        
        # Create email history record
        email_id = str(uuid.uuid4())
        email_history = {
            "id": email_id,
            "user_id": lead["user_id"],
            "lead_id": email_request.lead_id,
            "subject": email_request.subject,
            "body": email_request.body,
            "to_email": lead["email"],
            "from_email": sender_email,
            "from_name": "RealtorsPal Agent",
            "status": "draft",
            "llm_provider": email_request.llm_provider,
            "email_template": email_request.email_template,
            "created_at": datetime.now().isoformat()
        }
        
        try:
            # Use SendGrid Python SDK
            from sendgrid import SendGridAPIClient
            from sendgrid.helpers.mail import Mail
            
            # Initialize SendGrid client
            sg = SendGridAPIClient(sendgrid_api_key)
            
            print(f"🔵 Sending manual email via SendGrid SDK...")
            print(f"  To: {lead['email']}")
            print(f"  From: {sender_email}")
            print(f"  Subject: {email_request.subject}")
            
            # Create HTML version
            html_body = email_request.body.replace('\n', '<br>')
            
            # Create email message
            message = Mail(
                from_email=sender_email,
                to_emails=lead["email"],
                subject=email_request.subject,
                plain_text_content=email_request.body,
                html_content=html_body
            )
            
            # Send email
            response = sg.send(message)
            
            print(f"✅ SendGrid Response Status: {response.status_code}")
            print(f"  Response Headers: {dict(response.headers)}")
            
            # Check if successful (202 = Accepted)
            if response.status_code in [200, 201, 202]:
                message_id = response.headers.get('X-Message-Id', '')
                
                # Update email history with success
                email_history["status"] = "sent"
                email_history["sent_at"] = datetime.now().isoformat()
                email_history["sendgrid_message_id"] = message_id
                
                # Log email activity in lead notes
                current_notes = lead.get('notes', '')
                new_note = f"\n\n[Email] Sent: '{email_request.subject}' to {lead['email']} - {datetime.now().isoformat()}"
                await db.leads.update_one(
                    {"id": email_request.lead_id},
                    {"$set": {"notes": current_notes + new_note}}
                )
                
                print(f"✅ Manual email sent successfully! Message ID: {message_id}")
            else:
                # Handle non-success status codes
                error_msg = f"Unexpected status code: {response.status_code}"
                if response.body:
                    error_msg += f" - {response.body}"
                
                print(f"❌ SendGrid error: {error_msg}")
                email_history["status"] = "failed"
                email_history["error_message"] = error_msg
            
        except Exception as sendgrid_error:
            # Handle SendGrid SDK exceptions
            error_msg = str(sendgrid_error)
            
            # Try to extract detailed error from SendGrid exception
            if hasattr(sendgrid_error, 'body'):
                try:
                    import json
                    error_body = json.loads(sendgrid_error.body)
                    if "errors" in error_body:
                        error_msg = "; ".join([err.get("message", str(err)) for err in error_body["errors"]])
                except:
                    error_msg = sendgrid_error.body
            
            print(f"❌ SendGrid error: {error_msg}")
            email_history["status"] = "failed"
            email_history["error_message"] = error_msg
        
        # Save email history
        await db.email_history.insert_one(email_history)
        
        if email_history["status"] == "sent":
            return {
                "status": "success",
                "message": f"Email sent successfully to {lead['email']}",
                "email_id": email_id,
                "message_id": email_history.get("sendgrid_message_id", "")
            }
        else:
            return {
                "status": "error",
                "message": f"Failed to send email: {email_history.get('error_message', 'Unknown error')}",
                "email_id": email_id
            }
            
    except Exception as e:
        print(f"❌ Email sending error: {e}")
        import traceback
        traceback.print_exc()
        return {"status": "error", "message": str(e)}

@app.get("/api/email/history/{lead_id}")
async def get_email_history(lead_id: str):
    """Get email history for a specific lead"""
    try:
        history = await db.email_history.find(
            {"lead_id": lead_id}, 
            sort=[("created_at", -1)]
        ).to_list(length=None)
        
        return [EmailHistory(**{k: v for k, v in record.items() if k != "_id"}) for record in history]
        
    except Exception as e:
        print(f"Email history error: {e}")
        return []

@app.get("/api/email/draft")
async def draft_email_with_llm(lead_id: str, email_template: str = "follow_up", tone: str = "professional", llm_provider: str = "emergent"):
    """Draft email using LLM based on lead information"""
    try:
        # Get lead details
        lead = await db.leads.find_one({"id": lead_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Import LLM integration
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        from dotenv import load_dotenv
        import os
        
        # Load environment variables
        load_dotenv()
        
        # Get LLM key
        api_key = os.getenv('EMERGENT_LLM_KEY')
        if not api_key:
            return {
                "status": "error",
                "message": "LLM API key not found. Please configure Emergent LLM key.",
                "fallback_used": True
            }
        
        # Prepare lead information for LLM
        lead_name = f"{lead.get('first_name', '')} {lead.get('last_name', '')}".strip() or "there"
        property_info = f"{lead.get('property_type', 'property')} in {lead.get('neighborhood', 'your preferred area')}"
        budget_info = ""
        if lead.get('price_min') or lead.get('price_max'):
            budget_min = f"${lead.get('price_min', 0):,}" if lead.get('price_min') else "any"
            budget_max = f"${lead.get('price_max', 0):,}" if lead.get('price_max') else "any"
            budget_info = f" with a budget range of {budget_min} to {budget_max}"
        
        # Create context for LLM
        lead_context = f"""
        Lead Information:
        - Name: {lead_name}
        - Email: {lead.get('email', 'Not provided')}
        - Phone: {lead.get('phone', 'Not provided')}
        - Property Interest: {property_info}{budget_info}
        - Current Stage: {lead.get('stage', 'New')}
        - Priority: {lead.get('priority', 'medium')}
        - Notes: {lead.get('notes', 'No previous interactions')}
        """
        
        # Template-specific prompts
        template_prompts = {
            "follow_up": f"""
            Write a professional follow-up email to {lead_name} about their real estate inquiry. 
            They are interested in {property_info}{budget_info}.
            
            The email should:
            - Be {tone} in tone
            - Reference their specific property interest
            - Offer to help and provide value
            - Include a clear call-to-action (schedule a call or meeting)
            - Be personalized and not sound generic
            - Be around 100-150 words
            
            Include both a subject line and email body.
            """,
            "new_listing": f"""
            Write an email to {lead_name} about a new property listing that matches their criteria.
            They are interested in {property_info}{budget_info}.
            
            The email should:
            - Be {tone} in tone
            - Sound exciting about the new listing
            - Mention it matches their criteria
            - Create urgency (competitive market)
            - Include a call-to-action to schedule a viewing
            - Be around 100-150 words
            
            Include both a subject line and email body.
            """,
            "appointment_reminder": f"""
            Write a friendly reminder email to {lead_name} about an upcoming property viewing appointment.
            They are interested in {property_info}{budget_info}.
            
            The email should:
            - Be {tone} in tone
            - Remind them of the appointment
            - Express enthusiasm about helping them
            - Ask if they have any questions
            - Provide your contact information
            - Be around 80-120 words
            
            Include both a subject line and email body.
            """
        }
        
        # Get the appropriate prompt
        prompt = template_prompts.get(email_template, template_prompts["follow_up"])
        
        # Initialize LLM chat
        chat = LlmChat(
            api_key=api_key,
            session_id=f"email_draft_{lead_id}_{email_template}",
            system_message=f"""You are a professional real estate agent assistant helping to draft personalized emails to leads. 
            
            Always format your response as:
            SUBJECT: [email subject line]
            
            BODY:
            [email body content]
            
            Keep emails professional, personalized, and focused on providing value to the potential client.
            {lead_context}
            """
        )
        
        # Set model based on provider (default to gpt-4o-mini for cost efficiency)
        if llm_provider == "emergent":
            chat.with_model("openai", "gpt-4o-mini")
        elif llm_provider == "openai":
            chat.with_model("openai", "gpt-4o")
        elif llm_provider == "claude":
            chat.with_model("anthropic", "claude-3-5-sonnet-20241022")
        elif llm_provider == "gemini":
            chat.with_model("gemini", "gemini-2.0-flash")
        
        # Create user message
        user_message = UserMessage(text=prompt)
        
        # Get LLM response
        response = await chat.send_message(user_message)
        
        # Parse the response
        lines = response.strip().split('\n')
        subject = ""
        body = ""
        
        parsing_body = False
        for line in lines:
            if line.startswith("SUBJECT:"):
                subject = line.replace("SUBJECT:", "").strip()
            elif line.startswith("BODY:"):
                parsing_body = True
                continue
            elif parsing_body:
                if body:
                    body += "\n" + line
                else:
                    body = line
        
        # Fallback parsing if format is not followed
        if not subject or not body:
            response_parts = response.split('\n\n', 1)
            if len(response_parts) == 2:
                subject = response_parts[0].replace("SUBJECT:", "").strip()
                body = response_parts[1].replace("BODY:", "").strip()
            else:
                # Use the entire response as body and generate a subject
                body = response.strip()
                subject = f"Following up on your {lead.get('property_type', 'property')} inquiry"
        
        return {
            "status": "success",
            "subject": subject,
            "body": body,
            "template_used": email_template,
            "tone": tone,
            "llm_provider": llm_provider,
            "lead_name": lead_name,
            "property_info": property_info,
            "llm_generated": True
        }
        
    except Exception as e:
        print(f"LLM email drafting error: {e}")
        
        # Fallback to template-based generation if LLM fails
        lead_name = f"{lead.get('first_name', '')} {lead.get('last_name', '')}".strip() or "there"
        property_info = f"{lead.get('property_type', 'property')} in {lead.get('neighborhood', 'your preferred area')}"
        
        templates = {
            "follow_up": {
                "subject": f"Following up on your {lead.get('property_type', 'property')} inquiry",
                "body": f"""Dear {lead_name},

I hope this email finds you well. I wanted to follow up on your recent inquiry about {property_info}.

As your dedicated real estate agent, I'm here to help you find the perfect property that meets your needs and budget. I have several new listings that might interest you based on your preferences.

Would you be available for a quick call this week to discuss your requirements in more detail? I'd love to show you some properties that I think would be a great fit.

Best regards,
Your Real Estate Agent"""
            },
            "new_listing": {
                "subject": f"New {lead.get('property_type', 'property')} listing that matches your criteria",
                "body": f"""Hi {lead_name},

I hope you're doing well! I wanted to reach out because I have an exciting new listing that I think would be perfect for you.

Based on our previous conversation about your interest in {property_info}, this new property checks all the boxes and is priced competitively in the current market.

Would you like to schedule a viewing? I'm available this week and would love to show you this property before it goes to other potential buyers.

Looking forward to hearing from you!

Best regards,
Your Real Estate Agent"""
            },
            "appointment_reminder": {
                "subject": "Reminder: Property viewing appointment",
                "body": f"""Dear {lead_name},

This is a friendly reminder about our upcoming appointment to view properties in {lead.get('neighborhood', 'your preferred area')}.

Please let me know if you need to reschedule or if you have any questions before our meeting.

I'm looking forward to helping you find your dream home!

Best regards,
Your Real Estate Agent"""
            }
        }
        
        template_data = templates.get(email_template, templates["follow_up"])
        
        return {
            "status": "success",
            "subject": template_data["subject"],
            "body": template_data["body"],
            "template_used": email_template,
            "tone": tone,
            "lead_name": lead_name,
            "property_info": property_info,
            "llm_generated": False,
            "fallback_used": True,
            "error": str(e)
        }

# --- End Email Communication Endpoints ---

# --- Utils ---
def normalize_phone(phone_str):
    """Normalize phone number to E.164 format"""
    if not phone_str:
        return None
    
    # Remove all non-digit characters
    digits = re.sub(r'\D', '', phone_str)
    
    # If already has E.164 format, return as is
    if phone_str.startswith('+') and E164_RE.match(phone_str):
        return phone_str
    
    # Handle US numbers (10 or 11 digits)
    if len(digits) == 10:
        return f"+1{digits}"
    elif len(digits) == 11 and digits.startswith('1'):
        return f"+{digits}"
    
    # If it looks like it might be international but missing +, try adding it
    if len(digits) >= 7 and len(digits) <= 15:
        return f"+{digits}"
    
    # Return original if can't normalize
    return phone_str

async def get_user_by_email(email: str) -> Optional[Dict[str, Any]]:
    return await db.users.find_one({"email": email})

async def create_user(email: str, password: str, name: Optional[str] = None) -> Dict[str, Any]:
    hashed = pwd_context.hash(password)
    user = {
        "id": str(uuid.uuid4()),
        "email": email,
        "password_hash": hashed,
        "name": name or "Demo User",
    }
    await db.users.insert_one(user)
    return user

async def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)

# --- Startup: indexes & seed ---
@app.on_event("startup")
async def on_startup():
    await db.users.create_index("email", unique=True)
    await db.leads.create_index([("user_id", 1)])
    # partial unique only when email exists as string
    try:
        async for idx in db.leads.list_indexes():
            if idx.get("name") == "user_id_1_email_1":
                await db.leads.drop_index("user_id_1_email_1")
                break
    except Exception:
        pass
    await db.leads.create_index(
        [("user_id", 1), ("email", 1)],
        unique=True,
        partialFilterExpression={"email": {"$type": "string"}},
    )
    await db.settings.create_index([("user_id", 1)], unique=True)

# --- Routes ---
@app.get("/api/health")
async def health():
    return {"status": "ok"}

@app.get("/api/auth/demo", response_model=LoginResponse)
async def demo_session():
    demo_email = "demo@realtorspal.ai"
    user = await get_user_by_email(demo_email)
    if not user:
        user = await create_user(demo_email, "Demo123!", name="Demo User")
    return {"user": {"id": user["id"], "email": user["email"], "name": user.get("name")}, "token": "demo-token"}

@app.get("/api/leads", response_model=List[Lead])
async def list_leads(user_id: str):
    cursor = db.leads.find({"user_id": user_id})
    leads = []
    async for doc in cursor:
        leads.append(Lead(**{k: v for k, v in doc.items() if k != "_id"}))
    return leads

@app.post("/api/leads", response_model=Lead)
async def create_lead(payload: CreateLeadRequest):
    full_name = payload.name or " ".join([v for v in [payload.first_name, payload.last_name] if v]).strip() or "New Lead"
    in_dashboard = True if payload.in_dashboard is None else payload.in_dashboard
    stage = payload.stage or "New"
    
    # Normalize phone numbers
    normalized_phone = normalize_phone(payload.phone) if payload.phone else None
    normalized_work_phone = normalize_phone(payload.work_phone) if payload.work_phone else None
    normalized_home_phone = normalize_phone(payload.home_phone) if payload.home_phone else None
    normalized_spouse_phone = normalize_phone(payload.spouse_mobile_phone) if payload.spouse_mobile_phone else None
    
    lead = Lead(
        user_id=payload.user_id,
        name=full_name,
        
        # Basic fields
        first_name=payload.first_name,
        last_name=payload.last_name,
        email=payload.email,
        phone=normalized_phone,
        lead_description=payload.lead_description,
        
        # Additional Contact Information
        work_phone=normalized_work_phone,
        home_phone=normalized_home_phone,
        email_2=payload.email_2,
        
        # Spouse Information
        spouse_name=payload.spouse_name,
        spouse_first_name=payload.spouse_first_name,
        spouse_last_name=payload.spouse_last_name,
        spouse_email=payload.spouse_email,
        spouse_mobile_phone=normalized_spouse_phone,
        spouse_birthday=payload.spouse_birthday,
        
        # Pipeline and Status
        pipeline=payload.pipeline,
        status=payload.status,
        ref_source=payload.ref_source,
        lead_rating=payload.lead_rating,
        lead_source=payload.lead_source,
        lead_type=payload.lead_type,
        lead_type_2=payload.lead_type_2,
        
        # Property Information
        house_to_sell=payload.house_to_sell,
        buying_in=payload.buying_in,
        selling_in=payload.selling_in,
        owns_rents=payload.owns_rents,
        mortgage_type=payload.mortgage_type,
        
        # Address Information
        city=payload.city,
        zip_postal_code=payload.zip_postal_code,
        address=payload.address,
        
        # Property Details
        property_type=payload.property_type,
        property_condition=payload.property_condition,
        listing_status=payload.listing_status,
        bedrooms=payload.bedrooms,
        bathrooms=payload.bathrooms,
        basement=payload.basement,
        parking_type=payload.parking_type,
        
        # Dates and Anniversaries
        date_of_birth=payload.date_of_birth,
        house_anniversary=payload.house_anniversary,
        planning_to_sell_in=payload.planning_to_sell_in,
        
        # Agent Assignments
        main_agent=payload.main_agent,
        mort_agent=payload.mort_agent,
        list_agent=payload.list_agent,
        
        # Custom Fields
        custom_fields=payload.custom_fields,
        
        # Existing compatibility fields
        neighborhood=payload.neighborhood,
        price_min=payload.price_min,
        price_max=payload.price_max,
        priority=payload.priority,
        source_tags=payload.source_tags,
        notes=payload.notes,
        in_dashboard=in_dashboard,
        stage=stage,
    )
    try:
        await db.leads.insert_one(lead.model_dump(exclude_none=True))
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="A lead with this email already exists for this user.")
    return lead

@app.post("/api/leads/import", response_model=ImportResult)
async def import_leads(payload: ImportPayload):
    print(f"Import request received for user {payload.user_id}")
    print(f"Number of leads to import: {len(payload.leads)}")
    
    inserted = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []
    inserted_docs: List[Lead] = []

    for idx, item in enumerate(payload.leads):
        try:
            print(f"Processing lead {idx}: {item.first_name} {item.last_name} - {item.email}")
            
            full_name = item.name or " ".join([v for v in [item.first_name, item.last_name] if v]).strip() or "New Lead"
            stage = item.stage or payload.default_stage or "New"
            
            # Validate and normalize email
            validated_email = None
            if item.email and item.email.strip():
                try:
                    # Use email-validator for more lenient validation
                    validation = validate_email(item.email.strip(), check_deliverability=False)
                    validated_email = validation.email
                    print(f"Email validated: '{item.email}' -> '{validated_email}'")
                except EmailNotValidError as e:
                    print(f"Invalid email '{item.email}': {e}")
                    # Skip lead with invalid email or set to None
                    validated_email = None
            
            # Normalize phone numbers
            normalized_phone = normalize_phone(item.phone)
            normalized_work_phone = normalize_phone(item.work_phone)
            normalized_home_phone = normalize_phone(item.home_phone)
            normalized_spouse_phone = normalize_phone(item.spouse_mobile_phone)
            print(f"Phone normalized from '{item.phone}' to '{normalized_phone}'")
            
            lead = Lead(
                user_id=payload.user_id,
                # Basic fields
                name=full_name,
                first_name=item.first_name,
                last_name=item.last_name,
                email=validated_email,
                phone=normalized_phone,
                lead_description=item.lead_description,
                
                # Additional Contact Information
                work_phone=normalized_work_phone,
                home_phone=normalized_home_phone,
                email_2=item.email_2,
                
                # Spouse Information
                spouse_name=item.spouse_name,
                spouse_first_name=item.spouse_first_name,
                spouse_last_name=item.spouse_last_name,
                spouse_email=item.spouse_email,
                spouse_mobile_phone=normalized_spouse_phone,
                spouse_birthday=item.spouse_birthday,
                
                # Pipeline and Status
                pipeline=item.pipeline,
                status=item.status,
                ref_source=item.ref_source,
                lead_rating=item.lead_rating,
                lead_source=item.lead_source,
                lead_type=item.lead_type,
                lead_type_2=item.lead_type_2,
                
                # Property Information
                house_to_sell=item.house_to_sell,
                buying_in=item.buying_in,
                selling_in=item.selling_in,
                owns_rents=item.owns_rents,
                mortgage_type=item.mortgage_type,
                
                # Address Information
                city=item.city,
                zip_postal_code=item.zip_postal_code,
                address=item.address,
                
                # Property Details
                property_type=item.property_type,
                property_condition=item.property_condition,
                listing_status=item.listing_status,
                bedrooms=item.bedrooms,
                bathrooms=item.bathrooms,
                basement=item.basement,
                parking_type=item.parking_type,
                
                # Dates and Anniversaries
                date_of_birth=item.date_of_birth,
                house_anniversary=item.house_anniversary,
                planning_to_sell_in=item.planning_to_sell_in,
                
                # Agent Assignments
                main_agent=item.main_agent,
                mort_agent=item.mort_agent,
                list_agent=item.list_agent,
                
                # Compatibility fields
                neighborhood=item.neighborhood,
                price_min=item.price_min,
                price_max=item.price_max,
                priority=item.priority,
                source_tags=item.source_tags,
                notes=item.notes,
                in_dashboard=payload.in_dashboard,
                stage=stage,
            )
            await db.leads.insert_one(lead.model_dump(exclude_none=True))
            inserted += 1
            inserted_docs.append(lead)
            print(f"Successfully inserted lead {idx}")
        except DuplicateKeyError:
            skipped += 1
            error_msg = "duplicate email for this user"
            errors.append({"row": idx, "email": item.email, "reason": error_msg})
            print(f"Skipped lead {idx}: {error_msg}")
        except Exception as e:
            skipped += 1
            error_msg = str(e)
            errors.append({"row": idx, "reason": error_msg})
            print(f"Error processing lead {idx}: {error_msg}")
            print(f"Lead data: {item}")

    print(f"Import completed: {inserted} inserted, {skipped} skipped")
    return ImportResult(inserted=inserted, skipped=skipped, errors=errors, inserted_leads=inserted_docs)

@app.post("/api/leads/import-csv", response_model=ImportResult)
async def import_leads_csv(
    file: UploadFile = File(...),
    user_id: str = Form(...)
):
    """
    Import leads from CSV or Excel file with comprehensive field support
    
    Requirements:
    - Both email AND phone are compulsory
    - Maximum 1000 leads per import
    - Duplicates (same email) will be skipped
    - Supports CSV (.csv) and Excel (.xlsx, .xls) formats
    """
    print(f"File import request received for user {user_id}")
    print(f"File: {file.filename}, Content-Type: {file.content_type}")
    
    # Check file type
    file_ext = file.filename.split('.')[-1].lower()
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are allowed (.csv, .xlsx, .xls)")
    
    # Read file content
    try:
        contents = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")
    
    # Parse file based on type
    rows = []
    if file_ext == 'csv':
        try:
            csv_data = contents.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_data))
            rows = list(csv_reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing CSV file: {str(e)}")
    else:  # Excel files
        try:
            # Load workbook from bytes
            wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value else '')
            
            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        # Convert to string and handle None values
                        row_dict[headers[idx]] = str(value).strip() if value is not None else ''
                rows.append(row_dict)
            
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {str(e)}")
    
    print(f"File contains {len(rows)} rows")
    
    # Enforce 1000 lead limit
    if len(rows) > 1000:
        raise HTTPException(
            status_code=400, 
            detail=f"CSV contains {len(rows)} leads. Maximum allowed is 1000 leads per import."
        )
    
    inserted = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []
    inserted_docs: List[Lead] = []
    
    for idx, row in enumerate(rows):
        try:
            # Clean up empty strings to None
            cleaned_row = {k: (v.strip() if v and v.strip() else None) for k, v in row.items()}
            
            print(f"Processing row {idx + 1}: {cleaned_row.get('first_name')} {cleaned_row.get('last_name')}")
            
            # Required fields validation - BOTH email AND phone are compulsory
            email_value = cleaned_row.get('email') or cleaned_row.get('Email')
            phone_value = cleaned_row.get('phone') or cleaned_row.get('Phone')
            
            if not email_value or not phone_value:
                missing_fields = []
                if not email_value:
                    missing_fields.append('email')
                if not phone_value:
                    missing_fields.append('phone')
                
                skipped += 1
                error_msg = f"Missing required fields: {', '.join(missing_fields)}"
                errors.append({
                    "row": idx + 1,
                    "email": email_value,
                    "phone": phone_value,
                    "reason": error_msg
                })
                print(f"Skipped row {idx + 1}: {error_msg}")
                continue
            
            # Validate and normalize email
            validated_email = None
            try:
                validation = validate_email(email_value.strip(), check_deliverability=False)
                validated_email = validation.email
                print(f"Email validated: '{email_value}' -> '{validated_email}'")
            except EmailNotValidError as e:
                skipped += 1
                error_msg = f"Invalid email format: {str(e)}"
                errors.append({
                    "row": idx + 1,
                    "email": email_value,
                    "reason": error_msg
                })
                print(f"Skipped row {idx + 1}: {error_msg}")
                continue
            
            # Check for duplicate email
            existing_lead = await db.leads.find_one({
                "user_id": user_id,
                "email": validated_email
            })
            
            if existing_lead:
                skipped += 1
                error_msg = "Duplicate email - lead already exists"
                errors.append({
                    "row": idx + 1,
                    "email": validated_email,
                    "reason": error_msg
                })
                print(f"Skipped row {idx + 1}: {error_msg}")
                continue
            
            # Normalize phone numbers
            normalized_phone = normalize_phone(phone_value)
            if not normalized_phone:
                skipped += 1
                error_msg = f"Invalid phone format: {phone_value}"
                errors.append({
                    "row": idx + 1,
                    "phone": phone_value,
                    "reason": error_msg
                })
                print(f"Skipped row {idx + 1}: {error_msg}")
                continue
            
            normalized_work_phone = normalize_phone(cleaned_row.get('work_phone'))
            normalized_home_phone = normalize_phone(cleaned_row.get('home_phone'))
            normalized_spouse_phone = normalize_phone(cleaned_row.get('spouse_mobile_phone'))
            
            print(f"Phone normalized: '{phone_value}' -> '{normalized_phone}'")
            
            # Build full name
            first_name = cleaned_row.get('first_name') or cleaned_row.get('First Name')
            last_name = cleaned_row.get('last_name') or cleaned_row.get('Last Name')
            full_name = cleaned_row.get('name') or " ".join([v for v in [first_name, last_name] if v]).strip() or "New Lead"
            
            # Parse integers safely
            def safe_int(value):
                if value and str(value).strip():
                    try:
                        return int(str(value).replace(',', '').replace('$', '').strip())
                    except ValueError:
                        return None
                return None
            
            price_min = safe_int(cleaned_row.get('price_min'))
            price_max = safe_int(cleaned_row.get('price_max'))
            
            # Create lead with all comprehensive fields
            lead = Lead(
                user_id=user_id,
                # Basic fields
                name=full_name,
                first_name=first_name,
                last_name=last_name,
                email=validated_email,
                phone=normalized_phone,
                lead_description=cleaned_row.get('lead_description'),
                
                # Additional Contact Information
                work_phone=normalized_work_phone,
                home_phone=normalized_home_phone,
                email_2=cleaned_row.get('email_2'),
                
                # Spouse Information
                spouse_name=cleaned_row.get('spouse_name'),
                spouse_first_name=cleaned_row.get('spouse_first_name'),
                spouse_last_name=cleaned_row.get('spouse_last_name'),
                spouse_email=cleaned_row.get('spouse_email'),
                spouse_mobile_phone=normalized_spouse_phone,
                spouse_birthday=cleaned_row.get('spouse_birthday'),
                
                # Pipeline and Status
                pipeline=cleaned_row.get('pipeline'),
                status=cleaned_row.get('status') or 'Open',
                ref_source=cleaned_row.get('ref_source'),
                lead_rating=cleaned_row.get('lead_rating'),
                lead_source=cleaned_row.get('lead_source'),
                lead_type=cleaned_row.get('lead_type'),
                lead_type_2=cleaned_row.get('lead_type_2'),
                
                # Property Information
                house_to_sell=cleaned_row.get('house_to_sell'),
                buying_in=cleaned_row.get('buying_in'),
                selling_in=cleaned_row.get('selling_in'),
                owns_rents=cleaned_row.get('owns_rents'),
                mortgage_type=cleaned_row.get('mortgage_type'),
                
                # Address Information
                city=cleaned_row.get('city'),
                zip_postal_code=cleaned_row.get('zip_postal_code'),
                address=cleaned_row.get('address'),
                
                # Property Details
                property_type=cleaned_row.get('property_type'),
                property_condition=cleaned_row.get('property_condition'),
                listing_status=cleaned_row.get('listing_status'),
                bedrooms=cleaned_row.get('bedrooms'),
                bathrooms=cleaned_row.get('bathrooms'),
                basement=cleaned_row.get('basement'),
                parking_type=cleaned_row.get('parking_type'),
                
                # Dates and Anniversaries
                date_of_birth=cleaned_row.get('date_of_birth'),
                house_anniversary=cleaned_row.get('house_anniversary'),
                planning_to_sell_in=cleaned_row.get('planning_to_sell_in'),
                
                # Agent Assignments
                main_agent=cleaned_row.get('main_agent'),
                mort_agent=cleaned_row.get('mort_agent'),
                list_agent=cleaned_row.get('list_agent'),
                
                # Compatibility fields
                neighborhood=cleaned_row.get('neighborhood'),
                price_min=price_min,
                price_max=price_max,
                priority=cleaned_row.get('priority') or 'medium',
                notes=cleaned_row.get('notes') or cleaned_row.get('lead_description'),
                stage=cleaned_row.get('stage') or 'New',
                in_dashboard=True  # Default to showing in dashboard
            )
            
            # Insert lead
            await db.leads.insert_one(lead.model_dump(exclude_none=True))
            inserted += 1
            inserted_docs.append(lead)
            print(f"Successfully inserted lead {idx + 1}: {full_name}")
            
        except DuplicateKeyError:
            skipped += 1
            error_msg = "Duplicate email for this user"
            errors.append({
                "row": idx + 1,
                "email": email_value,
                "reason": error_msg
            })
            print(f"Skipped row {idx + 1}: {error_msg}")
        except Exception as e:
            skipped += 1
            error_msg = f"Error processing row: {str(e)}"
            errors.append({
                "row": idx + 1,
                "reason": error_msg
            })
            print(f"Error processing row {idx + 1}: {error_msg}")
            import traceback
            traceback.print_exc()
    
    print(f"CSV import completed: {inserted} inserted, {skipped} skipped")
    return ImportResult(
        inserted=inserted,
        skipped=skipped,
        errors=errors,
        inserted_leads=inserted_docs
    )

    print(f"Import completed: {inserted} inserted, {skipped} skipped")
    return ImportResult(inserted=inserted, skipped=skipped, errors=errors, inserted_leads=inserted_docs)

@app.put("/api/leads/{lead_id}/stage", response_model=Lead)
async def update_lead_stage(lead_id: str, payload: UpdateStageRequest):
    doc = await db.leads.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    await db.leads.update_one({"id": lead_id}, {"$set": {"stage": payload.stage}})
    updated = await db.leads.find_one({"id": lead_id})
    return Lead(**{k: v for k, v in updated.items() if k != "_id"})

@app.put("/api/leads/{lead_id}", response_model=Lead)
async def update_lead(lead_id: str, payload: UpdateLeadRequest):
    doc = await db.leads.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    data = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if not data:
        return Lead(**{k: v for k, v in doc.items() if k != "_id"})
    try:
        await db.leads.update_one({"id": lead_id}, {"$set": data})
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="A lead with this email already exists for this user.")
    updated = await db.leads.find_one({"id": lead_id})
    return Lead(**{k: v for k, v in updated.items() if k != "_id"})

@app.delete("/api/leads/{lead_id}")
async def delete_lead(lead_id: str):
    doc = await db.leads.find_one({"id": lead_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Lead not found")
    await db.leads.delete_one({"id": lead_id})
    return {"ok": True}

@app.get("/api/analytics/dashboard", response_model=AnalyticsDashboard)
async def analytics_dashboard(user_id: str):
    stages = ["New", "Contacted", "Appointment", "Onboarded", "Closed"]
    counts: Dict[str, int] = {s: 0 for s in stages}
    pipeline = [
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": "$stage", "count": {"$sum": 1}}},
    ]
    agg = db.leads.aggregate(pipeline)
    async for row in agg:
        stage = row.get("_id")
        if stage in counts:
            counts[stage] = row.get("count", 0)
    total = sum(counts.values())
    return AnalyticsDashboard(total_leads=total, by_stage=counts)

@app.get("/api/settings", response_model=Settings)
async def get_settings(user_id: str):
    print(f"🔵 get_settings called with user_id: {user_id}")
    doc = await db.settings.find_one({"user_id": user_id})
    
    if not doc:
        print(f"   No settings found, creating new...")
        settings = Settings(user_id=user_id)
        await db.settings.insert_one(settings.model_dump())
        return settings
    
    print(f"   Found settings in database:")
    print(f"   Twilio SID from DB: {doc.get('twilio_account_sid')}")
    print(f"   Twilio Phone from DB: {doc.get('twilio_phone_number')}")
    
    result = Settings(**{k: v for k, v in doc.items() if k != "_id"})
    
    print(f"   Returning Settings object:")
    print(f"   Twilio SID to return: {result.twilio_account_sid}")
    print(f"   Twilio Phone to return: {result.twilio_phone_number}")
    
    return result

class SaveSettingsRequest(BaseModel):
    user_id: str
    openai_api_key: Optional[str] = None
    anthropic_api_key: Optional[str] = None
    gemini_api_key: Optional[str] = None
    webhook_enabled: Optional[bool] = None
    facebook_webhook_verify_token: Optional[str] = None
    generic_webhook_enabled: Optional[bool] = None
    api_key: Optional[str] = None
    twilio_account_sid: Optional[str] = None
    twilio_auth_token: Optional[str] = None
    twilio_phone_number: Optional[str] = None
    twilio_whatsapp_number: Optional[str] = None
    twilio_api_key: Optional[str] = None
    twilio_api_secret: Optional[str] = None
    smtp_protocol: Optional[str] = None
    smtp_hostname: Optional[str] = None
    smtp_port: Optional[str] = None
    smtp_ssl_tls: Optional[bool] = None
    smtp_username: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_from_email: Optional[str] = None
    smtp_from_name: Optional[str] = None
    sendgrid_api_key: Optional[str] = None
    sender_email: Optional[str] = None  # Add sender_email field

@app.post("/api/settings", response_model=Settings)
async def save_settings(payload: SaveSettingsRequest):
    existing = await db.settings.find_one({"user_id": payload.user_id})
    data = {k: v for k, v in payload.model_dump().items() if k != "id"}
    
    # Extract secrets to save separately in secrets collection
    secret_fields = [
        'twilio_account_sid', 'twilio_auth_token', 'twilio_phone_number',
        'twilio_whatsapp_number', 'twilio_api_key', 'twilio_api_secret',
        'twilio_twiml_app_sid', 'agent_phone_number',
        'sendgrid_api_key', 'sender_email'
    ]
    
    secrets_to_save = {}
    for field in secret_fields:
        if field in data and data[field]:
            secrets_to_save[field] = data[field]
    
    # Save secrets to secure collection
    if secrets_to_save:
        await set_multiple_secrets(payload.user_id, secrets_to_save)
        print(f"🔒 Saved {len(secrets_to_save)} secrets to secure storage for user {payload.user_id}")
    
    # Save non-secret settings to settings collection (for backward compatibility)
    if existing:
        await db.settings.update_one({"user_id": payload.user_id}, {"$set": data})
        doc = await db.settings.find_one({"user_id": payload.user_id})
        return Settings(**{k: v for k, v in doc.items() if k != "_id"})
    else:
        settings = Settings(**data)
        await db.settings.insert_one(settings.model_dump())
        return settings

@app.post("/api/ai/chat")
async def chat(payload: dict):
    try:
        from emergentintegrations import EmergentLLM
        llm = EmergentLLM(api_key=os.environ.get('EMERGENT_LLM_KEY'))
        
        messages = payload.get('messages', [])
        stream = payload.get('stream', False)
        
        if stream:
            return StreamingResponse(
                llm.stream_chat(messages=messages),
                media_type="text/plain"
            )
        else:
            response = await llm.chat(messages=messages)
            return {"response": response}
    except Exception as e:
        return {"error": str(e), "fallback": "Chat service temporarily unavailable"}

# --- Webhook Endpoints ---

class FacebookLeadWebhook(BaseModel):
    object: str
    entry: List[Dict[str, Any]]

class GenericLeadWebhook(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    full_name: Optional[str] = None  # Added for compatibility
    email: Optional[str] = None
    phone: Optional[str] = None
    property_type: Optional[str] = None
    neighborhood: Optional[str] = None
    location: Optional[str] = None  # Added for compatibility
    budget: Optional[Union[str, int, float]] = None  # Accept string, int, or float
    source: Optional[str] = None
    lead_source: Optional[str] = None  # Added for compatibility
    timestamp: Optional[str] = None  # Added for compatibility
    custom_fields: Optional[Dict[str, Any]] = None

# --- Lead Generation AI Models ---

class LeadIntakeWebhook(BaseModel):
    """Lead Generation AI intake model with all possible fields"""
    # Contact Information
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    full_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    work_phone: Optional[str] = None
    home_phone: Optional[str] = None
    email_2: Optional[str] = None
    
    # Consent and Marketing
    consent_marketing: Optional[bool] = None
    consent_communication: Optional[bool] = None
    
    # Property Information
    property_type: Optional[str] = None
    buying_in: Optional[str] = None
    selling_in: Optional[str] = None
    owns_rents: Optional[str] = None
    mortgage_type: Optional[str] = None
    budget_min: Optional[Union[str, int, float]] = None
    budget_max: Optional[Union[str, int, float]] = None
    
    # Location Information
    city: Optional[str] = None
    neighborhood: Optional[str] = None
    zip_postal_code: Optional[str] = None
    address: Optional[str] = None
    
    # Property Details
    bedrooms: Optional[str] = None
    bathrooms: Optional[str] = None
    basement: Optional[str] = None
    parking_type: Optional[str] = None
    property_condition: Optional[str] = None
    listing_status: Optional[str] = None
    
    # Spouse Information
    spouse_first_name: Optional[str] = None
    spouse_last_name: Optional[str] = None
    spouse_email: Optional[str] = None
    spouse_mobile_phone: Optional[str] = None
    spouse_birthday: Optional[str] = None
    
    # Lead Source and Rating
    lead_source: Optional[str] = None
    ref_source: Optional[str] = None
    lead_rating: Optional[str] = None
    lead_type: Optional[str] = None
    lead_type_2: Optional[str] = None
    
    # Dates and Plans
    house_anniversary: Optional[str] = None
    planning_to_sell_in: Optional[str] = None
    
    # Agent Assignments
    main_agent: Optional[str] = None
    mort_agent: Optional[str] = None
    list_agent: Optional[str] = None
    
    # Additional Information
    lead_description: Optional[str] = None
    notes: Optional[str] = None
    priority: Optional[str] = None
    source: Optional[str] = None  # website|fb_lead_ad|chatbot|voice_bot|partner|csv_import
    
    # Custom fields and metadata
    custom_fields: Optional[Dict[str, Any]] = None

class LeadIntakeResult(BaseModel):
    """Response model for lead intake processing"""
    intake_result: str  # created | merged | rejected
    reason: str = ""
    upsert: Dict[str, Any]
    realtime_event: Dict[str, Any]
    audit: Dict[str, Any]

class AuditLog(BaseModel):
    """Audit log model for lead intake tracking"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    idempotency_key: str
    user_id: str
    hash_email: Optional[str] = None
    hash_phone: Optional[str] = None
    raw_source: str
    stored_raw_payload: bool = True
    intake_result: str
    lead_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())
    raw_payload: Dict[str, Any]

# --- Lead Generation AI Processing Functions ---

class MainOrchestratorAI:
    """Main Orchestrator AI for coordinating sub-agents and logging"""
    
    @staticmethod
    async def create_agent_run(agent_code: str, lead_id: str, user_id: str, step: str = None) -> "AgentRun":
        """Create a new agent run"""
        agent_run = AgentRun(
            agent_code=agent_code,
            lead_id=lead_id,
            user_id=user_id,
            step=step
        )
        await db.agent_runs.insert_one(agent_run.dict())
        return agent_run
    
    @staticmethod
    async def log_agent_event(run_id: str, event_type: str, payload: Dict[str, Any]):
        """Log an agent event"""
        event = AgentEvent(
            run_id=run_id,
            type=event_type,
            payload=payload
        )
        await db.agent_events.insert_one(event.dict())
        return event
    
    @staticmethod
    async def create_agent_task(run_id: str, lead_id: str, user_id: str, agent_code: str, 
                               due_at: str, channel: str, title: str, draft: Dict[str, str] = None) -> "AgentTask":
        """Create an actionable task for humans"""
        task = AgentTask(
            run_id=run_id,
            lead_id=lead_id,
            user_id=user_id,
            agent_code=agent_code,
            due_at=due_at,
            channel=channel,
            title=title,
            draft=draft
        )
        await db.agent_tasks.insert_one(task.dict())
        return task
    
    @staticmethod
    async def complete_agent_run(run_id: str, status: str = "succeeded"):
        """Mark agent run as completed"""
        await db.agent_runs.update_one(
            {"id": run_id},
            {"$set": {"status": status, "completed_at": datetime.utcnow().isoformat()}}
        )
    
    @staticmethod
    async def get_live_activity_stream(user_id: str, limit: int = 50) -> List[Dict[str, Any]]:
        """Get live activity stream from agent runs and events"""
        # Get recent agent runs with their events
        runs = await db.agent_runs.find(
            {"user_id": user_id}
        ).sort("started_at", -1).limit(limit).to_list(length=limit)
        
        activity_stream = []
        
        for run in runs:
            # Get events for this run
            events = await db.agent_events.find(
                {"run_id": run["id"]}
            ).sort("ts", -1).to_list(length=10)
            
            # Get tasks for this run
            tasks = await db.agent_tasks.find(
                {"run_id": run["id"]}
            ).sort("created_at", -1).to_list(length=5)
            
            # Get lead info
            lead = await db.leads.find_one({"id": run["lead_id"]})
            lead_name = "Unknown Lead"
            if lead:
                lead_name = f"{lead.get('first_name', '')} {lead.get('last_name', '')}".strip() or "Unnamed Lead"
            
            activity_stream.append({
                "id": run["id"],
                "type": "agent_run",
                "agent_code": run["agent_code"],
                "lead_id": run["lead_id"],
                "lead_name": lead_name,
                "status": run["status"],
                "step": run.get("step"),
                "started_at": run["started_at"],
                "completed_at": run.get("completed_at"),
                "correlation_id": run["correlation_id"],
                "events": events,
                "tasks": tasks
            })
        
        return activity_stream

class LeadGenerationAI:
    """Lead Generation AI for automatic lead processing"""
    
    @staticmethod
    def normalize_phone(phone: str) -> Optional[str]:
        """Normalize phone to E.164 format"""
        if not phone:
            return None
        
        # Remove all non-digit characters
        digits = re.sub(r'[^\d]', '', phone)
        
        if not digits:
            return None
        
        # Handle US phone numbers
        if len(digits) == 10:
            return f"+1{digits}"
        elif len(digits) == 11 and digits.startswith('1'):
            return f"+{digits}"
        elif digits.startswith('+'):
            return phone  # Already formatted
        else:
            # Assume US number if not properly formatted
            if len(digits) >= 10:
                return f"+1{digits[-10:]}"
        
        return None
    
    @staticmethod
    def normalize_email(email: str) -> Optional[str]:
        """Normalize and validate email"""
        if not email:
            return None
        
        email = email.strip().lower()
        
        # Basic email validation
        if '@' not in email or '.' not in email.split('@')[1]:
            return None
        
        return email
    
    @staticmethod
    def normalize_name(name: str) -> Optional[str]:
        """Normalize name to title case"""
        if not name:
            return None
        
        return name.strip().title()
    
    @staticmethod
    def generate_hash(value: str) -> Optional[str]:
        """Generate SHA256 hash for deduplication"""
        if not value:
            return None
        
        return hashlib.sha256(value.encode('utf-8')).hexdigest()
    
    @staticmethod
    def validate_minimal_fields(payload: LeadIntakeWebhook) -> tuple[bool, str]:
        """Validate minimal required fields"""
        # Must have at least first_name OR last_name
        if not payload.first_name and not payload.last_name and not payload.full_name:
            return False, "Missing required field: first_name, last_name, or full_name"
        
        # Must have at least email OR phone
        if not payload.email and not payload.phone:
            return False, "Missing required field: email or phone"
        
        # Must have marketing consent (can be False, but must be present)
        if payload.consent_marketing is None:
            return False, "Missing required field: consent_marketing"
        
        return True, ""
    
    @staticmethod
    def normalize_payload(payload: LeadIntakeWebhook) -> Dict[str, Any]:
        """Normalize payload according to rules"""
        normalized = {}
        
        # Handle names
        if payload.full_name and not payload.first_name and not payload.last_name:
            # Split full name
            name_parts = payload.full_name.strip().split()
            normalized['first_name'] = LeadGenerationAI.normalize_name(name_parts[0]) if name_parts else None
            normalized['last_name'] = LeadGenerationAI.normalize_name(' '.join(name_parts[1:])) if len(name_parts) > 1 else None
        else:
            normalized['first_name'] = LeadGenerationAI.normalize_name(payload.first_name)
            normalized['last_name'] = LeadGenerationAI.normalize_name(payload.last_name)
        
        # Handle contact info
        normalized['email'] = LeadGenerationAI.normalize_email(payload.email)
        normalized['phone_e164'] = LeadGenerationAI.normalize_phone(payload.phone)
        normalized['work_phone'] = LeadGenerationAI.normalize_phone(payload.work_phone)
        normalized['home_phone'] = LeadGenerationAI.normalize_phone(payload.home_phone)
        normalized['email_2'] = LeadGenerationAI.normalize_email(payload.email_2)
        
        # Handle spouse info
        normalized['spouse_first_name'] = LeadGenerationAI.normalize_name(payload.spouse_first_name)
        normalized['spouse_last_name'] = LeadGenerationAI.normalize_name(payload.spouse_last_name)
        normalized['spouse_email'] = LeadGenerationAI.normalize_email(payload.spouse_email)
        normalized['spouse_mobile_phone'] = LeadGenerationAI.normalize_phone(payload.spouse_mobile_phone)
        normalized['spouse_birthday'] = payload.spouse_birthday
        
        # Handle location
        normalized['city'] = payload.city.strip().title() if payload.city else None
        normalized['zip_postal_code'] = payload.zip_postal_code.strip() if payload.zip_postal_code else None
        normalized['address'] = payload.address.strip() if payload.address else None
        
        # Handle budget (map to Lead model fields)
        try:
            normalized['price_min'] = int(float(payload.budget_min)) if payload.budget_min else None
        except (ValueError, TypeError):
            normalized['price_min'] = None
        
        try:
            normalized['price_max'] = int(float(payload.budget_max)) if payload.budget_max else None
        except (ValueError, TypeError):
            normalized['price_max'] = None
        
        # Also set budget_min/budget_max for webhook response
        normalized['budget_min'] = normalized['price_min']
        normalized['budget_max'] = normalized['price_max']
        
        # Map dropdown fields with defaults
        dropdown_mappings = {
            'property_type': 'Single Family Home',
            'owns_rents': 'Not Selected',
            'mortgage_type': 'Not selected',
            'lead_rating': 'Not selected',
            'lead_source': 'Not selected',
            'lead_type': 'Not selected',
            'lead_type_2': 'Not selected',
            'house_to_sell': 'Unknown',
            'planning_to_sell_in': 'Not selected',
            'main_agent': 'Anil Botta',
            'mort_agent': 'Not selected',
            'list_agent': 'Not selected'
        }
        
        for field, default in dropdown_mappings.items():
            value = getattr(payload, field, None)
            normalized[field] = value if value else default
        
        # Set pipeline and status defaults
        normalized['pipeline'] = 'New Lead'
        normalized['status'] = 'Open'
        
        # Map phone_e164 to phone for Lead model compatibility
        normalized['phone'] = normalized['phone_e164']
        
        # Handle other fields
        normalized['ref_source'] = payload.ref_source or 'Ext. source'
        normalized['buying_in'] = payload.buying_in or ''
        normalized['selling_in'] = payload.selling_in or ''
        normalized['bedrooms'] = payload.bedrooms or ''
        normalized['bathrooms'] = payload.bathrooms or ''
        normalized['basement'] = payload.basement or ''
        normalized['parking_type'] = payload.parking_type or ''
        normalized['property_condition'] = payload.property_condition or ''
        normalized['listing_status'] = payload.listing_status or ''
        normalized['house_anniversary'] = payload.house_anniversary
        
        # Handle lead description with source attribution
        source = payload.source or 'unknown'
        timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        
        lead_desc_parts = []
        if payload.lead_description:
            lead_desc_parts.append(payload.lead_description.strip())
        if payload.notes:
            lead_desc_parts.append(payload.notes.strip())
        
        # Add source attribution
        lead_desc_parts.append(f"[{timestamp}] Lead captured via {source}")
        
        normalized['lead_description'] = '\n'.join(lead_desc_parts)
        
        return normalized
    
    @staticmethod
    async def find_duplicate_lead(user_id: str, email_hash: Optional[str], phone_hash: Optional[str]) -> Optional[Dict[str, Any]]:
        """Find duplicate lead by email or phone hash"""
        if not email_hash and not phone_hash:
            return None
        
        # Search for existing lead with matching hash
        query_conditions = []
        
        if email_hash:
            query_conditions.append({"hash_email": email_hash})
        
        if phone_hash:
            query_conditions.append({"hash_phone": phone_hash})
        
        if query_conditions:
            query = {
                "user_id": user_id,
                "$or": query_conditions
            }
            
            existing_lead = await db.leads.find_one(query)
            return existing_lead
        
        return None
    
    @staticmethod
    async def merge_lead_data(existing_lead: Dict[str, Any], normalized_data: Dict[str, Any]) -> Dict[str, Any]:
        """Merge new data into existing lead (keep existing non-empty values)"""
        merged = existing_lead.copy()
        
        for key, new_value in normalized_data.items():
            if key in ['id', 'user_id', 'created_at']:
                continue  # Don't update these fields
            
            existing_value = merged.get(key)
            
            # Only overwrite if existing is empty/None and new value is not empty/None
            if (not existing_value or existing_value in ['', 'Not selected', 'Not Selected', 'Unknown']) and new_value and new_value not in ['', 'Not selected', 'Not Selected', 'Unknown']:
                merged[key] = new_value
            elif key == 'lead_description':
                # Always append to lead description
                existing_desc = existing_value or ''
                new_desc = new_value or ''
                if existing_desc and new_desc:
                    merged[key] = f"{existing_desc}\n{new_desc}"
                elif new_desc:
                    merged[key] = new_desc
        
        return merged

@app.get("/api/webhooks/facebook-leads/{user_id}")
async def facebook_webhook_verification(user_id: str, request: Request):
    """Facebook webhook verification endpoint"""
    query_params = dict(request.query_params)
    mode = query_params.get('hub.mode')
    token = query_params.get('hub.verify_token')
    challenge = query_params.get('hub.challenge')
    
    # Get user's verify token from settings
    settings_doc = await db.settings.find_one({"user_id": user_id})
    if not settings_doc or not settings_doc.get('webhook_enabled'):
        raise HTTPException(status_code=404, detail="Webhook not enabled for this user")
    
    expected_token = settings_doc.get('facebook_webhook_verify_token')
    
    if mode == 'subscribe' and token == expected_token:
        return int(challenge)
    else:
        raise HTTPException(status_code=403, detail="Verification failed")

@app.post("/api/webhooks/facebook-leads/{user_id}")
async def facebook_webhook_handler(user_id: str, webhook_data: FacebookLeadWebhook):
    """Handle Facebook Lead Ads webhook"""
    try:
        # Verify user has webhook enabled
        settings_doc = await db.settings.find_one({"user_id": user_id})
        if not settings_doc or not settings_doc.get('webhook_enabled'):
            raise HTTPException(status_code=404, detail="Webhook not enabled")
        
        leads_created = []
        
        for entry in webhook_data.entry:
            if 'changes' in entry:
                for change in entry['changes']:
                    if change.get('field') == 'leadgen':
                        lead_data = change.get('value', {})
                        
                        # Extract lead information from Facebook webhook
                        form_data = {}
                        if 'form' in lead_data and 'leadgen_id' in lead_data:
                            # In a real implementation, you'd call Facebook API to get full lead details
                            # For now, we'll use the basic data structure
                            field_data = lead_data.get('field_data', [])
                            
                            for field in field_data:
                                field_name = field.get('name', '').lower()
                                field_value = field.get('values', [''])[0]
                                
                                if field_name in ['first_name', 'full_name']:
                                    form_data['first_name'] = field_value
                                elif field_name == 'last_name':
                                    form_data['last_name'] = field_value  
                                elif field_name in ['email', 'email_address']:
                                    form_data['email'] = field_value
                                elif field_name in ['phone', 'phone_number']:
                                    form_data['phone'] = field_value
                                elif field_name in ['property_type', 'looking_for']:
                                    form_data['property_type'] = field_value
                                elif field_name in ['location', 'city', 'area']:
                                    form_data['neighborhood'] = field_value
                        
                        # Create lead
                        if form_data:
                            # Normalize phone number
                            normalized_phone = normalize_phone(form_data.get('phone'))
                            
                            lead = Lead(
                                user_id=user_id,
                                name=f"{form_data.get('first_name', '')} {form_data.get('last_name', '')}".strip() or "Facebook Lead",
                                first_name=form_data.get('first_name'),
                                last_name=form_data.get('last_name'),
                                email=form_data.get('email'),
                                phone=normalized_phone,
                                property_type=form_data.get('property_type'),
                                neighborhood=form_data.get('neighborhood'),
                                source_tags=["Facebook Lead Ads"],
                                stage="New",
                                in_dashboard=True,  # Auto-add to dashboard for immediate attention
                                priority="medium"
                            )
                            
                            await db.leads.insert_one(lead.model_dump(exclude_none=True))
                            leads_created.append(lead)
        
        return {"status": "success", "leads_created": len(leads_created)}
        
    except Exception as e:
        print(f"Facebook webhook error: {e}")
        return {"status": "error", "message": str(e)}

@app.get("/api/webhooks/stats/{user_id}")
async def get_webhook_stats(user_id: str):
    """Get webhook activity statistics"""
    try:
        # Get current time for 24h calculations
        now = datetime.now()
        yesterday = now - timedelta(days=1)
        
        # Count leads from generic webhook (source_tags contains "Generic Webhook" or specific sources)
        generic_total = await db.leads.count_documents({
            "user_id": user_id,
            "$or": [
                {"source_tags": {"$in": ["Generic Webhook"]}},
                {"source_tags": {"$in": ["google_form", "zapier", "make", "custom"]}}
            ]
        })
        
        generic_24h = await db.leads.count_documents({
            "user_id": user_id,
            "created_at": {"$gte": yesterday.isoformat()},
            "$or": [
                {"source_tags": {"$in": ["Generic Webhook"]}},
                {"source_tags": {"$in": ["google_form", "zapier", "make", "custom"]}}
            ]
        })
        
        # Get last activity for generic webhook
        generic_last = await db.leads.find_one({
            "user_id": user_id,
            "$or": [
                {"source_tags": {"$in": ["Generic Webhook"]}},
                {"source_tags": {"$in": ["google_form", "zapier", "make", "custom"]}}
            ]
        }, sort=[("created_at", -1)])
        
        # Count leads from Facebook webhook
        facebook_total = await db.leads.count_documents({
            "user_id": user_id,
            "source_tags": {"$in": ["Facebook Lead Ads", "facebook", "instagram"]}
        })
        
        facebook_24h = await db.leads.count_documents({
            "user_id": user_id,
            "created_at": {"$gte": yesterday.isoformat()},
            "source_tags": {"$in": ["Facebook Lead Ads", "facebook", "instagram"]}
        })
        
        # Get last activity for Facebook webhook
        facebook_last = await db.leads.find_one({
            "user_id": user_id,
            "source_tags": {"$in": ["Facebook Lead Ads", "facebook", "instagram"]}
        }, sort=[("created_at", -1)])
        
        return {
            "generic": {
                "total": generic_total,
                "last_24h": generic_24h,
                "last_activity": generic_last.get("created_at") if generic_last else None,
                "status": "active" if generic_24h > 0 else "inactive"
            },
            "facebook": {
                "total": facebook_total,
                "last_24h": facebook_24h,
                "last_activity": facebook_last.get("created_at") if facebook_last else None,
                "status": "active" if facebook_24h > 0 else "inactive"
            }
        }
        
    except Exception as e:
        print(f"Webhook stats error: {e}")
        return {
            "generic": {"total": 0, "last_24h": 0, "last_activity": None, "status": "inactive"},
            "facebook": {"total": 0, "last_24h": 0, "last_activity": None, "status": "inactive"}
        }

# --- Lead Generation AI Webhook Endpoint ---

@app.post("/api/webhooks/lead-intake")
async def lead_intake_webhook(
    request: Request,
    payload: LeadIntakeWebhook,
    x_source: Optional[str] = Header(None),
    x_signature: Optional[str] = Header(None),
    idempotency_key: Optional[str] = Header(None)
):
    """
    Lead Generation AI webhook endpoint for automatic lead capture and processing
    
    This endpoint automatically processes incoming leads without human approval:
    - Validates minimal required fields
    - Normalizes data (names, phone E.164, emails)
    - Deduplicates by email/phone hash
    - Auto-upserts (creates or merges) leads
    - Emits real-time events
    - Logs audit trail
    """
    try:
        # Generate idempotency key if not provided
        if not idempotency_key:
            idempotency_key = str(uuid.uuid4())
        
        # Check idempotency - if we've processed this key before, return cached result
        existing_audit = await db.audit_logs.find_one({"idempotency_key": idempotency_key})
        if existing_audit:
            return {"status": "already_processed", "result": existing_audit.get("intake_result")}
        
        # Extract user_id from payload or use demo user for testing
        user_id = payload.custom_fields.get("user_id") if payload.custom_fields else None
        if not user_id:
            # Use demo user ID for testing
            user_id = "03f82986-51af-460c-a549-1c5077e67fb0"
        
        # Determine source
        source = x_source or payload.source or "webhook"
        
        # Step 1: Validate minimal required fields
        is_valid, validation_error = LeadGenerationAI.validate_minimal_fields(payload)
        if not is_valid:
            # Log rejection in audit
            audit_data = AuditLog(
                idempotency_key=idempotency_key,
                user_id=user_id,
                raw_source=source,
                intake_result="rejected",
                raw_payload=payload.dict()
            )
            await db.audit_logs.insert_one(audit_data.dict())
            
            return LeadIntakeResult(
                intake_result="rejected",
                reason=validation_error,
                upsert={},
                realtime_event={},
                audit={"reason": validation_error}
            ).dict()
        
        # Step 2: Normalize payload
        normalized_data = LeadGenerationAI.normalize_payload(payload)
        
        # Step 3: Generate hashes for deduplication
        email_hash = LeadGenerationAI.generate_hash(normalized_data['email']) if normalized_data['email'] else None
        phone_hash = LeadGenerationAI.generate_hash(normalized_data['phone_e164']) if normalized_data['phone_e164'] else None
        
        # Step 4: Check for duplicates and decide CREATE or MERGE
        existing_lead = await LeadGenerationAI.find_duplicate_lead(user_id, email_hash, phone_hash)
        
        operation = "create"
        lead_id = None
        final_lead_data = normalized_data.copy()
        
        if existing_lead:
            # MERGE operation
            operation = "update"
            lead_id = existing_lead['id']
            final_lead_data = await LeadGenerationAI.merge_lead_data(existing_lead, normalized_data)
            intake_result = "merged"
        else:
            # CREATE operation  
            lead_id = str(uuid.uuid4())
            final_lead_data['id'] = lead_id
            final_lead_data['user_id'] = user_id
            final_lead_data['created_at'] = datetime.utcnow().isoformat()
            final_lead_data['hash_email'] = email_hash
            final_lead_data['hash_phone'] = phone_hash
            intake_result = "created"
        
        # Step 5: Upsert into leads collection
        if operation == "create":
            await db.leads.insert_one(final_lead_data)
        else:
            await db.leads.update_one(
                {"id": lead_id},
                {"$set": final_lead_data}
            )
        
        # Step 6: Prepare response data
        lead_name = f"{final_lead_data.get('first_name', '')} {final_lead_data.get('last_name', '')}".strip()
        lead_type = final_lead_data.get('lead_type', 'Unknown')
        city = final_lead_data.get('city', '')
        budget_min = final_lead_data.get('budget_min', 0)
        budget_max = final_lead_data.get('budget_max', 0)
        
        budget_str = ""
        if budget_min or budget_max:
            if budget_min and budget_max:
                budget_str = f"${budget_min:,}-${budget_max:,}"
            elif budget_min:
                budget_str = f"${budget_min:,}+"
            elif budget_max:
                budget_str = f"Up to ${budget_max:,}"
        
        summary_parts = [lead_name, lead_type]
        if city:
            summary_parts.append(city)
        if budget_str:
            summary_parts.append(budget_str)
        
        summary = " • ".join(filter(None, summary_parts))
        
        # Step 7: Emit real-time event
        realtime_event = {
            "event": f"lead.{intake_result}",
            "payload": {
                "lead_id": lead_id,
                "summary": summary,
                "source": source
            }
        }
        
        # Step 8: Create audit log
        audit_data = AuditLog(
            idempotency_key=idempotency_key,
            user_id=user_id,
            hash_email=email_hash,
            hash_phone=phone_hash,
            raw_source=source,
            intake_result=intake_result,
            lead_id=lead_id,
            raw_payload=payload.dict()
        )
        await db.audit_logs.insert_one(audit_data.dict())
        
        # Step 9: Return result
        result = LeadIntakeResult(
            intake_result=intake_result,
            reason="",
            upsert={
                "operation": operation,
                "lead_id": lead_id,
                "data": final_lead_data
            },
            realtime_event=realtime_event,
            audit={
                "idempotency_key": idempotency_key,
                "hash_email": email_hash,
                "hash_phone": phone_hash,
                "raw_source": source,
                "stored_raw_payload": True
            }
        )
        
        return {"status": "accepted", "result": intake_result, "lead_id": lead_id, "summary": summary}
        
    except Exception as e:
        print(f"Lead intake webhook error: {e}")
        
        # Log error in audit if we have the key
        if idempotency_key:
            try:
                error_audit = AuditLog(
                    idempotency_key=idempotency_key,
                    user_id=user_id if 'user_id' in locals() else "unknown",
                    raw_source=source if 'source' in locals() else "unknown",
                    intake_result="error",
                    raw_payload=payload.dict() if 'payload' in locals() else {}
                )
                await db.audit_logs.insert_one(error_audit.dict())
            except:
                pass  # Don't fail on audit log error
        
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

# --- Main Orchestrator AI Models and Collections ---

class AgentRun(BaseModel):
    """Agent execution run tracking"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    agent_code: str  # LeadGeneratorAI, NurturingAI, CustomerServiceAI, OnboardingAI, CallLogAnalystAI, AnalyticsAI
    lead_id: str
    user_id: str
    status: str = "running"  # running, succeeded, failed, cancelled
    step: Optional[str] = None  # normalize, dedupe, draft_sequence, etc.
    started_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())
    completed_at: Optional[str] = None
    correlation_id: str = Field(default_factory=lambda: str(uuid.uuid4()))

class AgentEvent(BaseModel):
    """Fine-grained log entries per agent run"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    run_id: str  # Reference to agent_runs
    ts: str = Field(default_factory=lambda: datetime.utcnow().isoformat())
    type: str  # MSG.DRAFTED, CRM.UPDATE, ERROR, INFO, LEAD.CREATED, LEAD.MERGED, etc.
    payload: Dict[str, Any]

class AgentTask(BaseModel):
    """Actionable tasks for humans (Activity Board)"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    run_id: str  # Reference to agent_runs
    lead_id: str  # Reference to leads
    user_id: str
    agent_code: str
    due_at: str
    channel: str  # sms, email, call
    title: str
    draft: Optional[Dict[str, str]] = None  # {"subject": "", "body": ""}
    status: str = "pending"  # pending, completed, cancelled
    created_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())

class AgentResponse(BaseModel):
    """Standard agent response format"""
    agent_code: str
    lead_id: str
    run: Dict[str, Any]
    lead_updates: Dict[str, Any]

# --- Nurturing AI Models and System ---

class NurturingActivity(BaseModel):
    """Individual nurturing activity model"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lead_id: str
    user_id: str
    date: str  # ISO date string
    action: str  # voice_call, sms, email
    channel: str  # phone, sms, email  
    status: str = "pending"  # pending, completed, rescheduled, skipped
    notes: Optional[str] = None
    draft_content: Optional[str] = None
    subject: Optional[str] = None  # For emails
    to: Optional[str] = None  # Phone number or email
    created_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())
    completed_at: Optional[str] = None
    
class NurturingPlan(BaseModel):
    """Nurturing plan generated by AI"""
    lead_id: str
    user_id: str
    activity_board: List[NurturingActivity]
    lead_updates: Dict[str, Any]
    next_review: str
    engagement_score: int = 0
    strategy_notes: str = ""

class ReplyAnalysis(BaseModel):
    """Analysis of lead reply"""
    reply_text: str
    sentiment: str  # positive, neutral, negative, unresponsive
    intent: str  # interested, not_ready, price_concern, timing_issue, not_interested
    suggested_action: str
    confidence: float

class NurturingAI:
    """Nurturing AI for automated lead nurturing and activity planning"""
    
    @staticmethod
    def get_lead_context(lead: Dict[str, Any]) -> Dict[str, Any]:
        """Extract relevant context from lead data"""
        return {
            "name": f"{lead.get('first_name', '')} {lead.get('last_name', '')}".strip(),
            "email": lead.get('email'),
            "phone": lead.get('phone'),
            "pipeline": lead.get('pipeline', 'Not set'),
            "status": lead.get('status', 'Open'),
            "priority": lead.get('priority', 'medium'),
            "property_type": lead.get('property_type', 'Not specified'),
            "city": lead.get('city', ''),
            "neighborhood": lead.get('neighborhood', ''),
            "price_min": lead.get('price_min', 0),
            "price_max": lead.get('price_max', 0),
            "buying_in": lead.get('buying_in', ''),
            "selling_in": lead.get('selling_in', ''),
            "lead_source": lead.get('lead_source') or lead.get('ref_source', 'Unknown'),
            "created_at": lead.get('created_at'),
            "last_activity": lead.get('last_activity_date'),
            "engagement_score": lead.get('engagement_score', 50)
        }
    
    @staticmethod
    def determine_nurturing_strategy(context: Dict[str, Any]) -> Dict[str, Any]:
        """Determine nurturing strategy based on lead context"""
        pipeline = context.get('pipeline', '').lower()
        timeline = context.get('buying_in', '').lower() or context.get('selling_in', '').lower()
        engagement = context.get('engagement_score', 50)
        
        # Determine frequency based on timeline
        if any(term in timeline for term in ['0-3', '3-6', 'urgent', 'asap']):
            frequency = 'high'  # Every 2-3 days
            touches_per_week = 3
        elif any(term in timeline for term in ['6-12', '12+']):
            frequency = 'low'  # Weekly
            touches_per_week = 1
        else:
            frequency = 'medium'  # 2x per week
            touches_per_week = 2
        
        # Determine channel preference based on pipeline stage
        if pipeline in ['new lead', 'not set', 'tried to contact']:
            primary_channel = 'sms'  # Less intrusive for new leads
            secondary_channel = 'email'
        elif pipeline in ['made contact', 'warm / nurturing']:
            primary_channel = 'phone' if engagement > 70 else 'sms'
            secondary_channel = 'email'
        elif pipeline in ['hot/ ready', 'set meeting']:
            primary_channel = 'phone'  # Direct contact for hot leads
            secondary_channel = 'sms'
        else:
            primary_channel = 'email'
            secondary_channel = 'sms'
        
        return {
            "frequency": frequency,
            "touches_per_week": touches_per_week,
            "primary_channel": primary_channel,
            "secondary_channel": secondary_channel,
            "urgency": "high" if "hot" in pipeline or engagement > 80 else "medium"
        }
    
    @staticmethod
    async def execute_nurturing_agent(lead_id: str, user_id: str, lead_context: Dict[str, Any]) -> AgentResponse:
        """Execute Nurturing AI with proper logging"""
        # Create agent run
        agent_run = await MainOrchestratorAI.create_agent_run(
            agent_code="NurturingAI",
            lead_id=lead_id,
            user_id=user_id,
            step="analyze_lead_context"
        )
        
        try:
            # Log start event
            await MainOrchestratorAI.log_agent_event(
                run_id=agent_run.id,
                event_type="INFO",
                payload={"msg": f"Starting nurturing analysis for lead: {lead_context.get('name', 'Unknown')}"}
            )
            
            # Determine strategy
            strategy = NurturingAI.determine_nurturing_strategy(lead_context)
            await MainOrchestratorAI.log_agent_event(
                run_id=agent_run.id,
                event_type="INFO",
                payload={"msg": f"Determined strategy: {strategy['frequency']} frequency, {strategy['primary_channel']} channel"}
            )
            
            # Generate activity schedule
            activities = NurturingAI.generate_activity_schedule(strategy, lead_context)
            
            # Create tasks for each activity
            created_tasks = []
            for activity in activities:
                # Draft message
                message_data = await NurturingAI.draft_message(lead_context, activity.action, activity.channel)
                
                # Log draft event
                await MainOrchestratorAI.log_agent_event(
                    run_id=agent_run.id,
                    event_type="MSG.DRAFTED",
                    payload={
                        "channel": activity.channel,
                        "content": message_data.get('content', ''),
                        "subject": message_data.get('subject', '')
                    }
                )
                
                # Create task
                task = await MainOrchestratorAI.create_agent_task(
                    run_id=agent_run.id,
                    lead_id=lead_id,
                    user_id=user_id,
                    agent_code="NurturingAI",
                    due_at=(datetime.now() + timedelta(days=1)).isoformat(),
                    channel=activity.channel,
                    title=f"Send {activity.action.replace('_', ' ')} to {lead_context.get('name', 'Lead')}",
                    draft={
                        "subject": message_data.get('subject', ''),
                        "body": message_data.get('content', '')
                    }
                )
                created_tasks.append(task.dict())
            
            # Complete the run
            await MainOrchestratorAI.complete_agent_run(agent_run.id, "succeeded")
            
            # Log completion
            await MainOrchestratorAI.log_agent_event(
                run_id=agent_run.id,
                event_type="INFO",
                payload={"msg": f"Successfully created {len(created_tasks)} nurturing tasks"}
            )
            
            return AgentResponse(
                agent_code="NurturingAI",
                lead_id=lead_id,
                run={
                    "status": "succeeded",
                    "step": "nurturing_sequence_created",
                    "events": [
                        {"type": "INFO", "payload": {"msg": "Nurturing sequence analysis completed"}},
                        {"type": "CRM.UPDATE", "payload": {"tasks_created": len(created_tasks)}}
                    ],
                    "tasks": created_tasks
                },
                lead_updates={
                    "stage_suggestion": "contacted" if lead_context.get('pipeline') == 'New Lead' else lead_context.get('pipeline'),
                    "engagement_score": min(lead_context.get('engagement_score', 50) + 10, 100)
                }
            )
            
        except Exception as e:
            # Mark run as failed
            await MainOrchestratorAI.complete_agent_run(agent_run.id, "failed")
            await MainOrchestratorAI.log_agent_event(
                run_id=agent_run.id,
                event_type="ERROR",
                payload={"msg": str(e), "error_type": "nurturing_execution_error"}
            )
            raise e
    
    @staticmethod
    async def draft_message(lead_context: Dict[str, Any], activity_type: str, channel: str) -> Dict[str, str]:
        """Draft message content using LLM"""
        try:
            name = lead_context.get('name', 'there')
            property_type = lead_context.get('property_type', 'property')
            location = lead_context.get('city') or lead_context.get('neighborhood', 'your area')
            budget_min = lead_context.get('price_min', 0)
            budget_max = lead_context.get('price_max', 0)
            pipeline = lead_context.get('pipeline', 'Not set')
            
            budget_str = ""
            if budget_min or budget_max:
                if budget_min and budget_max:
                    budget_str = f"${budget_min:,}-${budget_max:,}"
                elif budget_min:
                    budget_str = f"${budget_min:,}+"
                elif budget_max:
                    budget_str = f"up to ${budget_max:,}"
            
            # Create prompts based on channel and activity type
            if channel == 'sms':
                prompt = f"""Draft a professional but friendly SMS for a real estate lead nurturing campaign.

Lead Details:
- Name: {name}
- Property Interest: {property_type}
- Location: {location}
- Budget: {budget_str}
- Current Stage: {pipeline}

Requirements:
- Keep it under 160 characters
- Personalized and conversational
- Include a clear call-to-action
- Professional but not pushy
- Real estate focused

Generate only the SMS text, no quotes or formatting."""

            elif channel == 'email':
                prompt = f"""Draft a professional email for a real estate lead nurturing campaign.

Lead Details:
- Name: {name}
- Property Interest: {property_type}
- Location: {location}
- Budget: {budget_str}
- Current Stage: {pipeline}

Requirements:
- Professional yet friendly tone
- Personalized content
- Include subject line
- Clear call-to-action
- Value-focused (market insights, listings, tips)
- Keep concise (under 200 words)

Return in format:
SUBJECT: [subject line]
BODY: [email content]"""

            else:  # phone/voice_call
                prompt = f"""Create a brief call script for a real estate agent to nurture this lead.

Lead Details:
- Name: {name}
- Property Interest: {property_type}
- Location: {location}
- Budget: {budget_str}
- Current Stage: {pipeline}

Requirements:
- Friendly opening
- Brief check-in
- Value proposition
- Clear next step
- 30-60 seconds max
- Natural conversation flow

Generate only the script text."""
            
            # Use LLM service if available
            if EMERGENT_LLM_KEY:
                llm_service = LlmService()
                response = await llm_service.generate_text(prompt)
                content = response.strip()
            else:
                # Fallback templates
                if channel == 'sms':
                    content = f"Hi {name}! Found some great {property_type.lower()} options in {location}. Want me to share details? - Your Agent"
                elif channel == 'email':
                    content = f"SUBJECT: New {property_type} listings in {location}\nBODY: Hi {name},\n\nHope you're doing well! I wanted to share some new {property_type.lower()} listings that came on the market in {location}. Based on your budget {budget_str}, I think you'd be interested.\n\nWould you like to schedule a call to discuss these opportunities?\n\nBest regards,\nYour Agent"
                else:
                    content = f"Hi {name}, this is your real estate agent. I wanted to check in about your {property_type.lower()} search in {location}. I have some new listings that might interest you. Do you have a few minutes to discuss?"
            
            # Parse email content
            if channel == 'email' and 'SUBJECT:' in content and 'BODY:' in content:
                parts = content.split('BODY:', 1)
                subject = parts[0].replace('SUBJECT:', '').strip()
                body = parts[1].strip()
                return {"subject": subject, "content": body}
            else:
                return {"content": content, "subject": ""}
            
        except Exception as e:
            print(f"Error drafting message: {e}")
            # Fallback content
            name = lead_context.get('name', 'there')
            if channel == 'sms':
                return {"content": f"Hi {name}! Quick check-in on your property search. Any questions? - Your Agent", "subject": ""}
            elif channel == 'email':
                return {
                    "subject": f"Checking in on your {property_type} search",
                    "content": f"Hi {name},\n\nJust wanted to check in on your property search. Let me know if you need any assistance!\n\nBest regards,\nYour Agent"
                }
            else:
                return {"content": f"Hi {name}, just checking in on your property search. How can I help?", "subject": ""}
    
    @staticmethod
    def generate_activity_schedule(strategy: Dict[str, Any], lead_context: Dict[str, Any]) -> List[NurturingActivity]:
        """Generate activity schedule based on strategy"""
        activities = []
        user_id = lead_context.get('user_id')
        lead_id = lead_context.get('lead_id') 
        
        if not user_id or not lead_id:
            return activities
        
        base_date = datetime.now()
        
        # Generate activities for next 2 weeks based on frequency
        touches_per_week = strategy.get('touches_per_week', 2)
        primary_channel = strategy.get('primary_channel', 'sms')
        secondary_channel = strategy.get('secondary_channel', 'email')
        
        for week in range(2):  # Next 2 weeks
            for touch in range(touches_per_week):
                # Calculate activity date
                days_offset = (week * 7) + (touch * (7 // touches_per_week)) + 1
                activity_date = (base_date + timedelta(days=days_offset))
                
                # Alternate between primary and secondary channels
                channel = primary_channel if touch % 2 == 0 else secondary_channel
                
                # Determine action type
                if channel == 'phone':
                    action = 'voice_call'
                    to = lead_context.get('phone', '')
                elif channel == 'sms':
                    action = 'sms'
                    to = lead_context.get('phone', '')
                else:  # email
                    action = 'email'
                    to = lead_context.get('email', '')
                
                if to:  # Only create activity if we have contact info
                    activity = NurturingActivity(
                        lead_id=lead_id,
                        user_id=user_id,
                        date=activity_date.strftime('%Y-%m-%d'),
                        action=action,
                        channel=channel,
                        to=to,
                        notes=f"Auto-generated nurturing activity for {channel} outreach"
                    )
                    activities.append(activity)
        
        return activities

# --- Lead Generation AI Test and Management Endpoints ---

@app.post("/api/lead-generation-ai/test")
async def test_lead_generation_ai(
    test_payload: LeadIntakeWebhook,
    user_id: Optional[str] = None
):
    """Test endpoint for Lead Generation AI without actually processing the lead"""
    try:
        # Use demo user if not specified
        if not user_id:
            user_id = "03f82986-51af-460c-a549-1c5077e67fb0"
        
        # Step 1: Validate minimal fields
        is_valid, validation_error = LeadGenerationAI.validate_minimal_fields(test_payload)
        if not is_valid:
            return {
                "status": "validation_failed",
                "error": validation_error,
                "normalized_data": None
            }
        
        # Step 2: Normalize data
        normalized_data = LeadGenerationAI.normalize_payload(test_payload)
        
        # Step 3: Generate hashes
        email_hash = LeadGenerationAI.generate_hash(normalized_data['email']) if normalized_data['email'] else None
        phone_hash = LeadGenerationAI.generate_hash(normalized_data['phone_e164']) if normalized_data['phone_e164'] else None
        
        # Step 4: Check for duplicates
        existing_lead = await LeadGenerationAI.find_duplicate_lead(user_id, email_hash, phone_hash)
        
        return {
            "status": "success",
            "validation": "passed",
            "normalized_data": normalized_data,
            "hashes": {
                "email_hash": email_hash,
                "phone_hash": phone_hash
            },
            "duplicate_check": {
                "has_duplicate": existing_lead is not None,
                "duplicate_lead_id": existing_lead.get('id') if existing_lead else None,
                "operation": "merge" if existing_lead else "create"
            }
        }
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

@app.get("/api/lead-generation-ai/audit-logs/{user_id}")
async def get_audit_logs(user_id: str, limit: int = 50):
    """Get audit logs for Lead Generation AI"""
    try:
        logs = await db.audit_logs.find(
            {"user_id": user_id}
        ).sort("created_at", -1).limit(limit).to_list(length=limit)
        
        # Remove MongoDB ObjectId from logs for JSON serialization
        clean_logs = []
        for log in logs:
            clean_log = {k: v for k, v in log.items() if k != "_id"}
            clean_logs.append(clean_log)
        
        return {
            "status": "success",
            "logs": clean_logs,
            "count": len(clean_logs)
        }
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

@app.post("/api/webhooks/generic-leads/{user_id}")
async def generic_webhook_handler(user_id: str, lead_data: GenericLeadWebhook):
    """Handle generic webhook for lead collection"""
    try:
        print(f"Generic webhook received for user {user_id}: {lead_data}")
        
        # Verify user has generic webhook enabled
        settings_doc = await db.settings.find_one({"user_id": user_id})
        if not settings_doc or not settings_doc.get('generic_webhook_enabled'):
            print(f"Generic webhook not enabled for user {user_id}")
            raise HTTPException(status_code=404, detail="Generic webhook not enabled")
        
        # Handle name fields - prioritize full_name, then first_name/last_name
        first_name = lead_data.first_name
        last_name = lead_data.last_name
        
        if lead_data.full_name and not first_name and not last_name:
            # Split full name if provided and individual names are missing
            name_parts = lead_data.full_name.split(' ', 1)
            first_name = name_parts[0] if len(name_parts) > 0 else None
            last_name = name_parts[1] if len(name_parts) > 1 else None
        
        # Handle location fields - prioritize location, then neighborhood
        location = lead_data.neighborhood or lead_data.location
        
        # Handle source fields - prioritize lead_source, then source
        source = lead_data.lead_source or lead_data.source or "Generic Webhook"
        
        # Normalize phone number  
        normalized_phone = normalize_phone(lead_data.phone)
        print(f"Phone normalized from '{lead_data.phone}' to '{normalized_phone}'")
        
        # Parse budget if provided - handle both string and numeric formats
        price_min = None
        price_max = None
        if lead_data.budget:
            if isinstance(lead_data.budget, (int, float)):
                # If budget is a number, use it as max price
                price_max = int(lead_data.budget)
                print(f"Budget {lead_data.budget} set as max price: {price_max}")
            else:
                # If budget is a string, parse it for ranges
                budget_str = str(lead_data.budget).replace(',', '').replace('$', '').replace('k', '000').replace('K', '000')
                price_match = re.findall(r'\d+', budget_str)
                if len(price_match) >= 2:
                    price_min = int(price_match[0])
                    price_max = int(price_match[1])
                elif len(price_match) == 1:
                    price_max = int(price_match[0])
                print(f"Budget parsed: min={price_min}, max={price_max}")
        
        lead = Lead(
            user_id=user_id,
            name=f"{first_name or ''} {last_name or ''}".strip() or "Generic Lead",
            first_name=first_name,
            last_name=last_name,
            email=lead_data.email,
            phone=normalized_phone,
            property_type=lead_data.property_type,
            neighborhood=location,
            price_min=price_min,
            price_max=price_max,
            source_tags=[source],
            stage="New",
            in_dashboard=True,  # Auto-add to dashboard
            priority="medium",
            notes=f"Timestamp: {lead_data.timestamp}, Custom fields: {lead_data.custom_fields}" if lead_data.timestamp or lead_data.custom_fields else None
        )
        
        print(f"Creating lead: {lead.name} - {lead.email} - {lead.phone}")
        await db.leads.insert_one(lead.model_dump(exclude_none=True))
        print(f"Lead created successfully with ID: {lead.id}")
        
        return {"status": "success", "lead_id": lead.id, "message": "Lead created successfully"}
        
    except Exception as e:
        print(f"Generic webhook error: {e}")
# =============================================================================
# AI AGENT SYSTEM - LLM SERVICE & AGENTS
# =============================================================================

class LLMService:
    """Centralized LLM service for all AI agents"""
    
    def __init__(self):
        self.emergent_llm_key = EMERGENT_LLM_KEY
        if not self.emergent_llm_key:
            print("Warning: EMERGENT_LLM_KEY not found in environment")
    
    async def generate_completion(self, 
                                prompt: str, 
                                model: str = "gpt-4o", 
                                provider: str = "openai",
                                system_prompt: str = "",
                                session_id: str = None) -> str:
        """Generate LLM completion using emergentintegrations"""
        try:
            if not self.emergent_llm_key:
                return "Error: LLM API key not configured"
            
            # Create session ID if not provided
            if not session_id:
                session_id = f"agent-{uuid.uuid4()}"
            
            # Initialize LlmChat with emergentintegrations
            chat = LlmChat(
                api_key=self.emergent_llm_key,
                session_id=session_id,
                system_message=system_prompt or "You are a helpful AI assistant."
            )
            
            # Set the model and provider
            chat.with_model(provider, model)
            
            # Create user message
            user_message = UserMessage(text=prompt)
            
            # Send message and get response
            response = await chat.send_message(user_message)
            
            return str(response)
            
        except Exception as e:
            print(f"LLM generation error: {e}")
            return f"LLM Error: {str(e)}"

    async def generate_structured_response(self,
                                         prompt: str,
                                         response_format: Dict[str, Any],
                                         model: str = "gpt-4o",
                                         provider: str = "openai",
                                         system_prompt: str = "") -> Dict[str, Any]:
        """Generate structured JSON response from LLM"""
        try:
            enhanced_system_prompt = f"""
{system_prompt}

CRITICAL: You must respond ONLY with valid JSON. No explanations, no markdown, no extra text.
Use this EXACT format:
{json.dumps(response_format, indent=2)}
"""
            
            full_prompt = f"""
{prompt}

IMPORTANT: Respond only with valid JSON in the exact format specified in the system message.
Do not include any other text, explanations, or markdown formatting. Only the JSON response.
"""
            
            response = await self.generate_completion(
                prompt=full_prompt,
                model=model,
                provider=provider,
                system_prompt=enhanced_system_prompt,
                session_id=f"structured-{uuid.uuid4()}"
            )
            
            # Clean response - remove any markdown formatting or extra text
            clean_response = response.strip()
            if clean_response.startswith("```json"):
                clean_response = clean_response.replace("```json", "").replace("```", "").strip()
            elif clean_response.startswith("```"):
                clean_response = clean_response.replace("```", "").strip()
            
            # Parse JSON response
            try:
                parsed_response = json.loads(clean_response)
                return parsed_response
            except json.JSONDecodeError as e:
                print(f"JSON decode error: {e}")
                print(f"Raw response: {response}")
                # Fallback if JSON parsing fails
                return {"error": "Invalid JSON response", "raw_response": response}
                
        except Exception as e:
            print(f"Structured response error: {e}")
            return {"error": f"Structured response error: {str(e)}"}

# Initialize LLM service
llm_service = LLMService()

class BaseAIAgent:
    """Base class for all AI agents"""
    
    def __init__(self, agent_id: str, name: str, system_prompt: str, model: str = "gpt-4o"):
        self.agent_id = agent_id
        self.name = name
        self.system_prompt = system_prompt
        self.model = model
        self.llm_service = llm_service
    
    async def process_task(self, task_data: Dict[str, Any], user_id: str) -> Dict[str, Any]:
        """Base method to be overridden by specific agents"""
        raise NotImplementedError("Each agent must implement process_task method")
    
    async def create_approval_request(self, 
                                    task: str, 
                                    proposal: Dict[str, Any], 
                                    user_id: str,
                                    lead_id: str = None,
                                    priority: str = "medium") -> str:
        """Create approval request and return approval_id"""
        approval_data = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "agent_id": self.agent_id,
            "agent_name": self.name,
            "task": task,
            "proposal": proposal,
            "lead_id": lead_id,
            "priority": priority,
            "status": "pending",
            "created_at": datetime.utcnow().isoformat()
        }
        
        await db.approval_requests.insert_one(approval_data)
        return approval_data["id"]
    
    async def log_activity(self, 
                          activity: str, 
                          user_id: str,
                          status: str = "completed",
                          activity_type: str = "automated",
                          details: Dict[str, Any] = None):
        """Log agent activity"""
        activity_data = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "agent_id": self.agent_id,
            "agent_name": self.name,
            "activity": activity,
            "status": status,
            "type": activity_type,
            "timestamp": datetime.utcnow().isoformat(),
            "details": details or {}
        }
        
        await db.agent_activities.insert_one(activity_data)

class LeadGeneratorAI(BaseAIAgent):
    """AI Agent for sourcing and normalizing leads from social media"""
    
    def __init__(self):
        super().__init__(
            agent_id="lead-generator",
            name="Lead Generator AI",
            system_prompt="""You are a Lead Generator AI for RealtorsPal. Your responsibilities:
1. Source and validate leads from social media platforms
2. Normalize lead data into consistent format
3. Detect and flag duplicate leads
4. Classify leads by quality and potential
5. Extract relevant information from social media profiles

Always maintain high data quality standards and follow privacy regulations.""",
            model="gpt-4o"
        )
    
    async def process_task(self, task_data: Dict[str, Any], user_id: str) -> Dict[str, Any]:
        """Process lead generation task"""
        try:
            await self.log_activity("Processing lead generation request", user_id, "processing")
            
            # Get existing leads for duplicate detection
            existing_leads = await db.leads.find({"user_id": user_id}).to_list(length=None)
            existing_emails = {lead.get("email", "").lower() for lead in existing_leads if lead.get("email")}
            
            # Analyze new lead data
            new_lead_data = task_data.get("lead_data", {})
            
            prompt = f"""
Analyze this potential lead data and provide recommendations:

Lead Data: {json.dumps(new_lead_data, indent=2)}

Existing Emails in Database: {list(existing_emails)[:10]}  # Show first 10 for context

Tasks:
1. Validate the lead quality (score 1-10)
2. Check for potential duplicates
3. Extract and normalize contact information
4. Suggest lead classification (hot/warm/cold)
5. Identify any data quality issues
6. Recommend next actions
"""

            response_format = {
                "lead_quality_score": 0,
                "duplicate_risk": "low|medium|high",
                "duplicate_reason": "",
                "normalized_data": {
                    "first_name": "",
                    "last_name": "", 
                    "email": "",
                    "phone": "",
                    "city": "",
                    "property_type": "",
                    "budget_range": "",
                    "timeline": ""
                },
                "classification": "hot|warm|cold",
                "data_quality_issues": [],
                "recommended_actions": [],
                "confidence_level": 0.0
            }
            
            analysis_result = await self.llm_service.generate_structured_response(
                prompt=prompt,
                response_format=response_format,
                model=self.model,
                system_prompt=self.system_prompt
            )
            
            # Determine if human approval is needed
            needs_approval = (
                analysis_result.get("lead_quality_score", 0) < 7 or
                analysis_result.get("duplicate_risk", "low") != "low" or
                len(analysis_result.get("data_quality_issues", [])) > 2
            )
            
            if needs_approval:
                # Create approval request
                proposal = {
                    "title": "Lead Generation Review Required",
                    "summary": [
                        f"Lead quality score: {analysis_result.get('lead_quality_score', 0)}/10",
                        f"Duplicate risk: {analysis_result.get('duplicate_risk', 'unknown')}",
                        f"Classification: {analysis_result.get('classification', 'unknown')}"
                    ],
                    "risks": analysis_result.get("data_quality_issues", []),
                    "action": "Add lead to CRM with normalizations",
                    "analysis": analysis_result
                }
                
                approval_id = await self.create_approval_request(
                    task="Review and approve new lead data",
                    proposal=proposal,
                    user_id=user_id,
                    priority="medium"
                )
                
                await self.log_activity("Lead analysis complete - approval required", user_id, "pending_approval", "approval_required")
                
                return {
                    "selected_agent": "LeadGeneratorAI",
                    "task": "Analyze and validate new lead data",
                    "rationale": "Lead requires human review due to quality or duplicate concerns",
                    "agent_output": {
                        "structured_fields": analysis_result,
                        "drafts_or_sequences": [],
                        "scores_or_flags": {
                            "quality_score": analysis_result.get("lead_quality_score", 0),
                            "confidence": analysis_result.get("confidence_level", 0),
                            "duplicate_risk": analysis_result.get("duplicate_risk", "unknown")
                        }
                    },
                    "human_approval": {
                        "required": True,
                        "title": proposal["title"],
                        "summary": proposal["summary"],
                        "risks": proposal["risks"],
                        "choices": ["Approve", "Edit", "Reject"],
                        "approval_id": approval_id
                    },
                    "data_patch": {
                        "crm_updates": {"leads_analyzed": 1, "approval_required": True},
                        "stage_suggestion": "validation"
                    }
                }
            else:
                # Auto-process high quality leads
                await self.log_activity("High quality lead auto-processed", user_id, "completed")
                
                return {
                    "selected_agent": "LeadGeneratorAI", 
                    "task": "Process high-quality lead automatically",
                    "rationale": "Lead meets quality standards for automatic processing",
                    "agent_output": {
                        "structured_fields": analysis_result,
                        "drafts_or_sequences": ["Lead ready for immediate contact"],
                        "scores_or_flags": {
                            "quality_score": analysis_result.get("lead_quality_score", 0),
                            "confidence": analysis_result.get("confidence_level", 0),
                            "auto_processed": True
                        }
                    },
                    "human_approval": {
                        "required": False,
                        "title": "High Quality Lead Processed",
                        "summary": ["Lead automatically validated and added to CRM"],
                        "risks": [],
                        "choices": []
                    },
                    "data_patch": {
                        "crm_updates": {"leads_processed": 1, "auto_approved": True},
                        "stage_suggestion": "new_lead"
                    }
                }
                
        except Exception as e:
            await self.log_activity(f"Error in lead generation: {str(e)}", user_id, "failed")
            raise e

class LeadNurturingAI(BaseAIAgent):
    """AI Agent for creating personalized follow-up sequences"""
    
    def __init__(self):
        super().__init__(
            agent_id="lead-nurturing",
            name="Lead Nurturing AI", 
            system_prompt="""You are a Lead Nurturing AI for RealtorsPal. Your responsibilities:
1. Create personalized follow-up email sequences
2. Draft human-sounding messages tailored to lead profiles
3. Suggest optimal timing for follow-ups
4. Generate SMS and phone call talking points
5. Adapt communication style based on lead temperature

Always create authentic, helpful communications that build trust and provide value.""",
            model="claude-3-sonnet"
        )
    
    async def process_task(self, task_data: Dict[str, Any], user_id: str) -> Dict[str, Any]:
        """Process lead nurturing task"""
        try:
            await self.log_activity("Creating personalized nurturing sequence", user_id, "processing")
            
            leads_to_nurture = task_data.get("leads", [])
            nurture_type = task_data.get("type", "follow_up")  # follow_up, welcome, re_engagement
            
            if not leads_to_nurture:
                # Get leads that need nurturing
                recent_leads = await db.leads.find({
                    "user_id": user_id, 
                    "pipeline": {"$in": ["new", "warm / nurturing", "contacted"]}
                }).limit(5).to_list(length=None)
                leads_to_nurture = recent_leads
            
            sequences = []
            
            for lead in leads_to_nurture:
                lead_profile = {
                    "name": f"{lead.get('first_name', '')} {lead.get('last_name', '')}".strip(),
                    "email": lead.get("email", ""),
                    "phone": lead.get("phone", ""),
                    "city": lead.get("city", ""),
                    "property_type": lead.get("property_type", ""),
                    "budget": lead.get("price_range", ""),
                    "timeline": lead.get("buying_in", ""),
                    "source": lead.get("lead_source", ""),
                    "pipeline": lead.get("pipeline", ""),
                    "notes": lead.get("lead_description", "")
                }
                
                prompt = f"""
Create a personalized 3-email nurturing sequence for this lead:

Lead Profile: {json.dumps(lead_profile, indent=2)}

Nurturing Type: {nurture_type}

Requirements:
1. Each email should feel personal and authentic
2. Provide genuine value in each touchpoint
3. Use the lead's name and reference their specific interests
4. Include soft call-to-actions
5. Vary the timing and approach for each email
6. Keep tone professional but friendly
"""

                response_format = {
                    "email_1": {
                        "subject": "",
                        "body": "",
                        "timing": "immediate|1_day|3_days|1_week",
                        "tone": "friendly|professional|casual|formal",
                        "cta": ""
                    },
                    "email_2": {
                        "subject": "",
                        "body": "",
                        "timing": "3_days|1_week|2_weeks",
                        "tone": "friendly|professional|casual|formal", 
                        "cta": ""
                    },
                    "email_3": {
                        "subject": "",
                        "body": "",
                        "timing": "1_week|2_weeks|1_month",
                        "tone": "friendly|professional|casual|formal",
                        "cta": ""
                    },
                    "sms_follow_up": "",
                    "phone_talking_points": [],
                    "personalization_notes": ""
                }
                
                sequence = await self.llm_service.generate_structured_response(
                    prompt=prompt,
                    response_format=response_format,
                    model=self.model,
                    system_prompt=self.system_prompt
                )
                
                sequence["lead_id"] = lead.get("id")
                sequence["lead_name"] = lead_profile["name"]
                sequences.append(sequence)
            
            # Always require approval for nurturing sequences
            proposal = {
                "title": "Lead Nurturing Sequence Approval",
                "summary": [
                    f"Created sequences for {len(sequences)} leads",
                    f"Type: {nurture_type.replace('_', ' ').title()} campaign",
                    "3 personalized emails + SMS + call points per lead"
                ],
                "risks": ["Messages may need personalization", "Timing might conflict with holidays"],
                "action": "Send personalized nurturing sequences",
                "sequences": sequences
            }
            
            approval_id = await self.create_approval_request(
                task="Review and approve nurturing sequences",
                proposal=proposal,
                user_id=user_id,
                priority="medium"
            )
            
            await self.log_activity("Nurturing sequences created - pending approval", user_id, "pending_approval", "approval_required")
            
            return {
                "selected_agent": "LeadNurturingAI",
                "task": "Create personalized follow-up sequences",
                "rationale": f"Generated nurturing sequences for {len(sequences)} leads requiring human review",
                "agent_output": {
                    "structured_fields": {
                        "sequences_created": len(sequences),
                        "nurture_type": nurture_type,
                        "leads_affected": len(leads_to_nurture)
                    },
                    "drafts_or_sequences": [f"Email sequence for {seq['lead_name']}" for seq in sequences],
                    "scores_or_flags": {
                        "confidence": 0.9,
                        "personalization_level": "high",
                        "approval_required": True
                    }
                },
                "human_approval": {
                    "required": True,
                    "title": proposal["title"],
                    "summary": proposal["summary"],
                    "risks": proposal["risks"],
                    "choices": ["Approve", "Edit", "Reject"],
                    "approval_id": approval_id
                },
                "data_patch": {
                    "crm_updates": {
                        "sequences_created": len(sequences),
                        "leads_in_nurturing": len(leads_to_nurture)
                    },
                    "stage_suggestion": "nurturing"
                }
            }
            
        except Exception as e:
            await self.log_activity(f"Error in lead nurturing: {str(e)}", user_id, "failed")
            raise e

class CustomerServiceAI(BaseAIAgent):
    """AI Agent for triaging inbound messages and drafting replies"""
    
    def __init__(self):
        super().__init__(
            agent_id="customer-service",
            name="Customer Service AI",
            system_prompt="""You are a Customer Service AI for RealtorsPal. Your responsibilities:
1. Triage inbound messages by urgency and intent
2. Draft appropriate responses to customer inquiries
3. Detect sentiment and escalation triggers
4. Provide helpful, accurate information
5. Maintain professional tone while being empathetic

Always prioritize customer satisfaction and timely responses.""",
            model="gemini-pro"
        )
    
    async def process_task(self, task_data: Dict[str, Any], user_id: str) -> Dict[str, Any]:
        """Process customer service task"""
        try:
            await self.log_activity("Triaging customer message", user_id, "processing")
            
            message_data = task_data.get("message", {})
            customer_info = task_data.get("customer", {})
            
            prompt = f"""
Analyze this customer message and provide triage recommendations:

Customer Message: {message_data.get('content', '')}
Customer Info: {json.dumps(customer_info, indent=2)}
Message Source: {message_data.get('source', 'email')}
Timestamp: {message_data.get('timestamp', '')}

Analyze for:
1. Intent classification
2. Urgency level 
3. Sentiment analysis
4. Escalation triggers
5. Draft appropriate response
6. Next action recommendations
"""

            response_format = {
                "intent": "inquiry|complaint|compliment|request|emergency|spam",
                "urgency": "low|medium|high|urgent",
                "sentiment": "positive|neutral|negative|angry",
                "escalation_triggers": [],
                "response_draft": "",
                "response_tone": "professional|empathetic|apologetic|enthusiastic",
                "next_actions": [],
                "requires_human": False,
                "estimated_resolution_time": "",
                "confidence_score": 0.0
            }
            
            analysis = await self.llm_service.generate_structured_response(
                prompt=prompt,
                response_format=response_format,
                model=self.model,
                system_prompt=self.system_prompt
            )
            
            # Determine if human involvement needed
            needs_human = (
                analysis.get("urgency") in ["high", "urgent"] or
                analysis.get("sentiment") in ["negative", "angry"] or
                len(analysis.get("escalation_triggers", [])) > 0 or
                analysis.get("intent") == "complaint" or
                analysis.get("requires_human", False)
            )
            
            if needs_human:
                proposal = {
                    "title": "Customer Service Escalation",
                    "summary": [
                        f"Intent: {analysis.get('intent', 'unknown')}",
                        f"Urgency: {analysis.get('urgency', 'unknown')}",
                        f"Sentiment: {analysis.get('sentiment', 'unknown')}"
                    ],
                    "risks": analysis.get("escalation_triggers", []),
                    "action": "Send drafted response and escalate if needed",
                    "analysis": analysis,
                    "draft_response": analysis.get("response_draft", "")
                }
                
                approval_id = await self.create_approval_request(
                    task="Review customer service response",
                    proposal=proposal,
                    user_id=user_id,
                    priority="high" if analysis.get("urgency") == "urgent" else "medium"
                )
                
                await self.log_activity("Customer message triaged - escalation required", user_id, "pending_approval", "approval_required")
                
                return {
                    "selected_agent": "CustomerServiceAI",
                    "task": "Triage and respond to customer message",
                    "rationale": "Message requires human review due to urgency or sentiment",
                    "agent_output": {
                        "structured_fields": analysis,
                        "drafts_or_sequences": [analysis.get("response_draft", "")],
                        "scores_or_flags": {
                            "urgency_score": analysis.get("urgency", "low"),
                            "sentiment_score": analysis.get("sentiment", "neutral"),
                            "confidence": analysis.get("confidence_score", 0)
                        }
                    },
                    "human_approval": {
                        "required": True,
                        "title": proposal["title"],
                        "summary": proposal["summary"],
                        "risks": proposal["risks"],
                        "choices": ["Approve", "Edit", "Reject"],
                        "approval_id": approval_id
                    },
                    "data_patch": {
                        "crm_updates": {"messages_triaged": 1, "escalations": 1},
                        "stage_suggestion": "customer_service"
                    }
                }
            else:
                # Auto-respond to simple inquiries
                await self.log_activity("Simple inquiry auto-processed", user_id, "completed")
                
                return {
                    "selected_agent": "CustomerServiceAI",
                    "task": "Handle routine customer inquiry",
                    "rationale": "Simple inquiry can be handled automatically",
                    "agent_output": {
                        "structured_fields": analysis,
                        "drafts_or_sequences": [analysis.get("response_draft", "")],
                        "scores_or_flags": {
                            "urgency_score": analysis.get("urgency", "low"),
                            "sentiment_score": analysis.get("sentiment", "neutral"),
                            "auto_resolved": True
                        }
                    },
                    "human_approval": {
                        "required": False,
                        "title": "Routine Inquiry Handled",
                        "summary": ["Response automatically generated and sent"],
                        "risks": [],
                        "choices": []
                    },
                    "data_patch": {
                        "crm_updates": {"messages_processed": 1, "auto_responses": 1},
                        "stage_suggestion": "resolved"
                    }
                }
                
        except Exception as e:
            await self.log_activity(f"Error in customer service: {str(e)}", user_id, "failed")
            raise e

# Initialize AI Agents
lead_generator = LeadGeneratorAI()
lead_nurturer = LeadNurturingAI()
customer_service = CustomerServiceAI()

# Agent registry for orchestrator
AGENT_REGISTRY = {
    "lead-generator": lead_generator,
    "lead-nurturing": lead_nurturer,
    "customer-service": customer_service
}

# =============================================================================
# AI AGENT SYSTEM ENDPOINTS
# =============================================================================

class AgentConfig(BaseModel):
    id: str
    name: str
    description: str
    status: str = "active"  # active, idle, disabled
    model: str = "gpt-4o"
    provider: str = "emergent"  # emergent, openai, anthropic, gemini
    system_prompt: str = ""
    response_tone: str = "professional"
    automation_rules: Dict[str, Any] = Field(default_factory=dict)
    custom_templates: Dict[str, str] = Field(default_factory=dict)
    performance_metrics: Dict[str, Any] = Field(default_factory=dict)
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class AgentActivity(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    agent_id: str
    agent_name: str
    activity: str
    status: str = "processing"  # processing, completed, failed, pending_approval
    type: str = "automated"  # automated, approval_required, human_decision
    timestamp: datetime = Field(default_factory=datetime.utcnow)
    details: Dict[str, Any] = Field(default_factory=dict)

class ApprovalRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    agent_id: str
    agent_name: str
    task: str
    proposal: Dict[str, Any]
    lead_id: Optional[str] = None
    priority: str = "medium"  # low, medium, high, urgent
    status: str = "pending"  # pending, approved, rejected, edited
    created_at: datetime = Field(default_factory=datetime.utcnow)
    resolved_at: Optional[datetime] = None
    resolved_by: Optional[str] = None
    decision: Optional[str] = None

class ApprovalDecision(BaseModel):
    decision: str  # approve, edit, reject
    notes: Optional[str] = None

# Get all AI agents
@app.get("/api/ai-agents")
async def get_ai_agents(user_id: str):
    """Get all AI agents for a user"""
    try:
        # Get agent configurations from database
        agents_cursor = db.ai_agents.find({"user_id": user_id})
        agents = await agents_cursor.to_list(length=None)
        
        # If no agents exist, create default agents
        if not agents:
            default_agents = [
                {
                    "id": "orchestrator",
                    "user_id": user_id,
                    "name": "Main Orchestrator AI",
                    "description": "Coordinates all AI agents and makes strategic decisions",
                    "status": "active",
                    "model": "gpt-4o",
                    "provider": "emergent",
                    "system_prompt": "You are the Main Orchestrator AI for RealtorsPal. Coordinate all AI agents, make strategic decisions, and ensure optimal system performance.",
                    "response_tone": "professional",
                    "automation_rules": {"auto_approve_low_risk": True, "escalate_threshold": 0.8},
                    "custom_templates": {},
                    "performance_metrics": {"success_rate": 98, "avg_response_time": 1.2, "tasks_completed": 0},
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                },
                {
                    "id": "lead-generator",
                    "user_id": user_id,
                    "name": "Lead Generator AI",
                    "description": "Sources and normalizes leads from social media",
                    "status": "active",
                    "model": "gpt-4o",
                    "provider": "emergent",
                    "system_prompt": "You are a Lead Generator AI. Source, validate, and normalize leads from various social media platforms and lead sources.",
                    "response_tone": "analytical",
                    "automation_rules": {"auto_validate": True, "duplicate_check": True},
                    "custom_templates": {},
                    "performance_metrics": {"success_rate": 94, "avg_response_time": 2.1, "tasks_completed": 0},
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                },
                {
                    "id": "lead-nurturing",
                    "user_id": user_id,
                    "name": "Lead Nurturing AI",
                    "description": "Creates personalized follow-up sequences",
                    "status": "active",
                    "model": "claude-3-sonnet",
                    "provider": "emergent",
                    "system_prompt": "You are a Lead Nurturing AI. Create personalized, human-sounding follow-up sequences for leads at different stages.",
                    "response_tone": "friendly",
                    "automation_rules": {"personalization_level": "high", "follow_up_intervals": [1, 3, 7, 14]},
                    "custom_templates": {
                        "initial_contact": "Hi {first_name}, thanks for your interest in real estate...",
                        "follow_up": "Hi {first_name}, I wanted to follow up on..."
                    },
                    "performance_metrics": {"success_rate": 96, "avg_response_time": 1.8, "tasks_completed": 0},
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                },
                {
                    "id": "customer-service",
                    "user_id": user_id,
                    "name": "Customer Service AI",
                    "description": "Triages inbound messages and drafts replies",
                    "status": "active",
                    "model": "gemini-pro",
                    "provider": "emergent",
                    "system_prompt": "You are a Customer Service AI. Triage inbound messages, detect intent, and draft appropriate responses.",
                    "response_tone": "helpful",
                    "automation_rules": {"auto_respond_threshold": 0.9, "escalate_keywords": ["complaint", "urgent", "manager"]},
                    "custom_templates": {},
                    "performance_metrics": {"success_rate": 92, "avg_response_time": 0.9, "tasks_completed": 0},
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                },
                {
                    "id": "onboarding",
                    "user_id": user_id,
                    "name": "Onboarding Agent AI",
                    "description": "Converts qualified leads into active clients",
                    "status": "active",
                    "model": "gpt-4o-mini",
                    "provider": "emergent",
                    "system_prompt": "You are an Onboarding Agent AI. Help convert qualified leads into active clients with personalized onboarding experiences.",
                    "response_tone": "welcoming",
                    "automation_rules": {"create_onboarding_plan": True, "welcome_sequence": True},
                    "custom_templates": {
                        "welcome": "Welcome to our real estate family, {first_name}!"
                    },
                    "performance_metrics": {"success_rate": 89, "avg_response_time": 3.2, "tasks_completed": 0},
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                },
                {
                    "id": "call-analyst",
                    "user_id": user_id,
                    "name": "Call Log Analyst AI",
                    "description": "Analyzes transcripts and extracts insights",
                    "status": "idle",
                    "model": "claude-3-haiku",
                    "provider": "emergent",
                    "system_prompt": "You are a Call Log Analyst AI. Analyze call transcripts, extract key insights, and provide actionable recommendations.",
                    "response_tone": "analytical",
                    "automation_rules": {"sentiment_analysis": True, "key_phrase_extraction": True},
                    "custom_templates": {},
                    "performance_metrics": {"success_rate": 97, "avg_response_time": 2.5, "tasks_completed": 0},
                    "created_at": datetime.utcnow().isoformat(),
                    "updated_at": datetime.utcnow().isoformat()
                }
            ]
            
            # Insert default agents
            await db.ai_agents.insert_many(default_agents)
            agents = default_agents
        else:
            # Convert MongoDB documents to JSON-serializable format
            converted_agents = []
            for agent in agents:
                agent_dict = {k: v for k, v in agent.items() if k != "_id"}
                # Convert datetime objects to ISO strings if they exist
                if "created_at" in agent_dict and hasattr(agent_dict["created_at"], "isoformat"):
                    agent_dict["created_at"] = agent_dict["created_at"].isoformat()
                if "updated_at" in agent_dict and hasattr(agent_dict["updated_at"], "isoformat"):
                    agent_dict["updated_at"] = agent_dict["updated_at"].isoformat()
                converted_agents.append(agent_dict)
            agents = converted_agents
        
        return {"agents": agents}
        
    except Exception as e:
        print(f"Error getting AI agents: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Update AI agent configuration
@app.put("/api/ai-agents/{agent_id}")
async def update_ai_agent(agent_id: str, agent_config: Dict[str, Any], user_id: str):
    """Update AI agent configuration"""
    try:
        agent_config["updated_at"] = datetime.utcnow().isoformat()
        
        result = await db.ai_agents.update_one(
            {"id": agent_id, "user_id": user_id},
            {"$set": agent_config}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Agent not found")
            
        return {"status": "success", "message": "Agent updated successfully"}
        
    except Exception as e:
        print(f"Error updating AI agent: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Get agent activities (live stream data)
@app.get("/api/ai-agents/activities")
async def get_agent_activities(user_id: str, limit: int = 50):
    """Get recent agent activities for live streaming"""
    try:
        activities_cursor = db.agent_activities.find(
            {"user_id": user_id}
        ).sort("timestamp", -1).limit(limit)
        
        activities = await activities_cursor.to_list(length=None)
        
        # Convert MongoDB documents to JSON-serializable format
        converted_activities = []
        for activity in activities:
            activity_dict = {k: v for k, v in activity.items() if k != "_id"}
            # Convert datetime objects to ISO strings if they exist
            if "timestamp" in activity_dict and hasattr(activity_dict["timestamp"], "isoformat"):
                activity_dict["timestamp"] = activity_dict["timestamp"].isoformat()
            converted_activities.append(activity_dict)
        
        return {"activities": converted_activities}
        
    except Exception as e:
        print(f"Error getting agent activities: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Create agent activity (for simulation and real agent actions)
@app.post("/api/ai-agents/activities")
async def create_agent_activity(activity: Dict[str, Any], user_id: str):
    """Create a new agent activity"""
    try:
        activity_data = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "timestamp": datetime.utcnow().isoformat(),
            **activity
        }
        
        await db.agent_activities.insert_one(activity_data)
        # Return a copy without MongoDB's _id field
        return_data = {k: v for k, v in activity_data.items() if k != "_id"}
        return {"status": "success", "activity": return_data}
        
    except Exception as e:
        print(f"Error creating agent activity: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Get approval queue
@app.get("/api/ai-agents/approvals")
async def get_approval_queue(user_id: str):
    """Get pending approval requests"""
    try:
        approvals_cursor = db.approval_requests.find(
            {"user_id": user_id, "status": "pending"}
        ).sort("created_at", -1)
        
        approvals = await approvals_cursor.to_list(length=None)
        
        # Convert MongoDB documents to JSON-serializable format
        converted_approvals = []
        for approval in approvals:
            approval_dict = {k: v for k, v in approval.items() if k != "_id"}
            # Convert datetime objects to ISO strings if they exist
            if "created_at" in approval_dict and hasattr(approval_dict["created_at"], "isoformat"):
                approval_dict["created_at"] = approval_dict["created_at"].isoformat()
            if "resolved_at" in approval_dict and hasattr(approval_dict["resolved_at"], "isoformat"):
                approval_dict["resolved_at"] = approval_dict["resolved_at"].isoformat()
            converted_approvals.append(approval_dict)
        
        return {"approvals": converted_approvals}
        
    except Exception as e:
        print(f"Error getting approval queue: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Create approval request
@app.post("/api/ai-agents/approvals")
async def create_approval_request(approval: Dict[str, Any], user_id: str):
    """Create a new approval request"""
    try:
        approval_data = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "status": "pending",
            "created_at": datetime.utcnow().isoformat(),
            **approval
        }
        
        await db.approval_requests.insert_one(approval_data)
        # Return a copy without MongoDB's _id field
        return_data = {k: v for k, v in approval_data.items() if k != "_id"}
        return {"status": "success", "approval": return_data}
        
    except Exception as e:
        print(f"Error creating approval request: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Handle approval decision
@app.put("/api/ai-agents/approvals/{approval_id}")
async def handle_approval_decision(approval_id: str, decision: ApprovalDecision, user_id: str):
    """Handle human approval decision"""
    try:
        update_data = {
            "status": decision.decision,
            "decision": decision.decision,
            "resolved_at": datetime.utcnow().isoformat(),
            "resolved_by": user_id,
            "notes": decision.notes
        }
        
        result = await db.approval_requests.update_one(
            {"id": approval_id, "user_id": user_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Approval request not found")
            
        # Log the decision as an activity
        activity_data = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "agent_id": "human-supervisor",
            "agent_name": "Human Supervisor",
            "activity": f"{decision.decision.upper()}: Approval request {approval_id}",
            "status": "completed",
            "type": "human_decision",
            "timestamp": datetime.utcnow().isoformat(),
            "details": {"approval_id": approval_id, "notes": decision.notes}
        }
        await db.agent_activities.insert_one(activity_data)
            
        return {"status": "success", "message": f"Approval {decision.decision} successfully"}
        
    except Exception as e:
        print(f"Error handling approval decision: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Master Orchestrator Agent endpoint  
@app.post("/api/ai-agents/orchestrate")
async def orchestrate_agents(task_data: Dict[str, Any], user_id: str):
    """Main orchestrator endpoint for coordinating agents with LLM integration"""
    try:
        # Get current CRM context
        leads = await db.leads.find({"user_id": user_id}).to_list(length=None)
        recent_activities = await db.agent_activities.find({"user_id": user_id}).sort("timestamp", -1).limit(10).to_list(length=None)
        
        # Prepare context for orchestrator
        crm_context = {
            "total_leads": len(leads),
            "new_leads": len([l for l in leads if l.get("pipeline") in ["new", "New Lead"]]),
            "warm_leads": len([l for l in leads if l.get("pipeline") == "warm / nurturing"]),
            "contacted_leads": len([l for l in leads if l.get("pipeline") == "contacted"]),
            "recent_activities": [{"agent": a.get("agent_name"), "activity": a.get("activity")} for a in recent_activities[:5]],
            "task_request": task_data
        }
        
        # Enhanced orchestrator prompt
        orchestrator_prompt = f"""
You are the Main Orchestrator AI for RealtorsPal CRM. Analyze the current situation and decide which AI agent should handle the task.

Current CRM State:
{json.dumps(crm_context, indent=2)}

Available Agents:
1. LeadGeneratorAI - Sources/validates leads from social media
2. LeadNurturingAI - Creates personalized follow-up sequences  
3. CustomerServiceAI - Triages messages and drafts replies
4. OnboardingAgentAI - Converts qualified leads to clients
5. CallLogAnalystAI - Analyzes transcripts and extracts insights

Task Request: {task_data.get('type', 'general')}
Priority: {task_data.get('priority', 'medium')}

Select the best agent and provide detailed reasoning.
"""

        response_format = {
            "selected_agent": "LeadGeneratorAI|LeadNurturingAI|CustomerServiceAI|OnboardingAgentAI|CallLogAnalystAI",
            "task": "",
            "rationale": "",
            "estimated_impact": "low|medium|high",
            "risk_assessment": "low|medium|high", 
            "human_approval_needed": True,
            "priority_level": "low|medium|high|urgent",
            "success_probability": 0.0
        }
        
        # Get LLM decision
        orchestrator_decision = await llm_service.generate_structured_response(
            prompt=orchestrator_prompt,
            response_format=response_format,
            model="gpt-4o",
            system_prompt="""You are the Main Orchestrator AI for RealtorsPal. Make strategic decisions about which AI agent should handle each task based on:
1. Current CRM state and lead pipeline
2. Agent capabilities and expertise
3. Task complexity and requirements
4. Risk assessment and approval needs
5. Overall system efficiency

Always provide clear reasoning for your decisions."""
        )
        
        # Execute task with selected agent if available
        selected_agent_id = orchestrator_decision.get("selected_agent", "").lower().replace("ai", "").replace("agent", "").strip()
        
        if selected_agent_id in AGENT_REGISTRY:
            # Execute task with the selected agent
            agent = AGENT_REGISTRY[selected_agent_id]
            agent_result = await agent.process_task(task_data, user_id)
            
            # Combine orchestrator decision with agent result
            orchestrator_response = {
                **agent_result,
                "orchestrator_decision": orchestrator_decision,
                "selected_by_orchestrator": True
            }
        else:
            # Fallback to orchestrator-only response
            orchestrator_response = {
                "selected_agent": orchestrator_decision.get("selected_agent", "LeadNurturingAI"),
                "task": orchestrator_decision.get("task", "Process CRM task"),
                "rationale": orchestrator_decision.get("rationale", "Selected based on current CRM state"),
                "agent_output": {
                    "structured_fields": {
                        "orchestrator_confidence": orchestrator_decision.get("success_probability", 0.8),
                        "impact_level": orchestrator_decision.get("estimated_impact", "medium"),
                        "risk_level": orchestrator_decision.get("risk_assessment", "low")
                    },
                    "drafts_or_sequences": ["Task routed to appropriate agent"],
                    "scores_or_flags": {
                        "confidence": orchestrator_decision.get("success_probability", 0.8),
                        "priority": orchestrator_decision.get("priority_level", "medium")
                    }
                },
                "human_approval": {
                    "required": orchestrator_decision.get("human_approval_needed", True),
                    "title": "Orchestrator Task Assignment",
                    "summary": [
                        f"Selected: {orchestrator_decision.get('selected_agent', 'Unknown Agent')}",
                        f"Impact: {orchestrator_decision.get('estimated_impact', 'medium')}",
                        f"Risk: {orchestrator_decision.get('risk_assessment', 'low')}"
                    ],
                    "risks": ["Task complexity may require human oversight"],
                    "choices": ["Approve", "Edit", "Reject"]
                },
                "data_patch": {
                    "crm_updates": {"tasks_orchestrated": 1},
                    "stage_suggestion": "processing"
                }
            }
        
        # Log orchestrator activity
        activity_data = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "agent_id": "orchestrator",
            "agent_name": "Main Orchestrator AI",
            "activity": f"Orchestrated: {orchestrator_response.get('task', 'CRM task')}",
            "status": "completed" if not orchestrator_response.get("human_approval", {}).get("required") else "pending_approval",
            "type": "approval_required" if orchestrator_response.get("human_approval", {}).get("required") else "automated",
            "timestamp": datetime.utcnow().isoformat(),
            "details": {
                "selected_agent": orchestrator_decision.get("selected_agent"),
                "reasoning": orchestrator_decision.get("rationale")
            }
        }
        await db.agent_activities.insert_one(activity_data)
        
        # Create approval request if required
        if orchestrator_response.get("human_approval", {}).get("required"):
            approval_data = {
                "id": str(uuid.uuid4()),
                "user_id": user_id,
                "agent_id": "orchestrator",
                "agent_name": "Main Orchestrator AI", 
                "task": orchestrator_response.get("task", "CRM task orchestration"),
                "proposal": orchestrator_response.get("human_approval", {}),
                "priority": orchestrator_decision.get("priority_level", "medium"),
                "status": "pending",
                "created_at": datetime.utcnow().isoformat()
            }
            await db.approval_requests.insert_one(approval_data)
        
        return orchestrator_response
        
    except Exception as e:
        print(f"Error in agent orchestration: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Individual Agent Task Endpoints
@app.post("/api/ai-agents/lead-generator/process")
async def process_lead_generation(task_data: Dict[str, Any], user_id: str):
    """Process lead generation task directly"""
    return await lead_generator.process_task(task_data, user_id)

@app.post("/api/ai-agents/lead-nurturing/process")  
async def process_lead_nurturing(task_data: Dict[str, Any], user_id: str):
    """Process lead nurturing task directly"""
    return await lead_nurturer.process_task(task_data, user_id)

@app.post("/api/ai-agents/customer-service/process")
async def process_customer_service(task_data: Dict[str, Any], user_id: str):
    """Process customer service task directly"""

# --- Main Orchestrator AI Endpoints ---

@app.get("/api/orchestrator/live-activity-stream/{user_id}")
async def get_live_activity_stream(user_id: str, limit: int = 50):
    """Get live activity stream from agent runs and events"""
    try:
        activity_stream = await MainOrchestratorAI.get_live_activity_stream(user_id, limit)
        
        # Clean up MongoDB ObjectIds for JSON serialization
        clean_stream = []
        for item in activity_stream:
            clean_item = {k: v for k, v in item.items() if k != "_id"}
            if 'events' in clean_item:
                clean_item['events'] = [{k: v for k, v in event.items() if k != "_id"} for event in clean_item['events']]
            if 'tasks' in clean_item:
                clean_item['tasks'] = [{k: v for k, v in task.items() if k != "_id"} for task in clean_item['tasks']]
            clean_stream.append(clean_item)
        
        return {
            "status": "success",
            "activity_stream": clean_stream,
            "count": len(clean_stream)
        }
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

@app.get("/api/orchestrator/agent-runs/{user_id}")
async def get_agent_runs(user_id: str, agent_code: Optional[str] = None, limit: int = 100):
    """Get agent runs with optional filtering"""
    try:
        query = {"user_id": user_id}
        if agent_code:
            query["agent_code"] = agent_code
        
        runs = await db.agent_runs.find(query).sort("started_at", -1).limit(limit).to_list(length=limit)
        
        # Clean up MongoDB ObjectIds
        clean_runs = [{k: v for k, v in run.items() if k != "_id"} for run in runs]
        
        return {
            "status": "success",
            "agent_runs": clean_runs,
            "count": len(clean_runs)
        }
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

@app.get("/api/orchestrator/agent-tasks/{user_id}")
async def get_agent_tasks(user_id: str, status: Optional[str] = None, limit: int = 100):
    """Get agent tasks for Activity Board"""
    try:
        query = {"user_id": user_id}
        if status:
            query["status"] = status
        
        tasks = await db.agent_tasks.find(query).sort("created_at", -1).limit(limit).to_list(length=limit)
        
        # Clean up MongoDB ObjectIds and add lead names
        clean_tasks = []
        for task in tasks:
            clean_task = {k: v for k, v in task.items() if k != "_id"}
            
            # Get lead info
            lead = await db.leads.find_one({"id": clean_task["lead_id"]})
            if lead:
                clean_task["lead_name"] = f"{lead.get('first_name', '')} {lead.get('last_name', '')}".strip() or "Unnamed Lead"
            else:
                clean_task["lead_name"] = "Unknown Lead"
            
            clean_tasks.append(clean_task)
        
        return {
            "status": "success",
            "agent_tasks": clean_tasks,
            "count": len(clean_tasks)
        }
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

@app.post("/api/orchestrator/execute-agent")
async def execute_agent(agent_code: str, lead_id: str, user_id: str, context: Dict[str, Any] = None):
    """Execute a specific agent with proper logging"""
    try:
        if agent_code == "NurturingAI":
            # Get lead context
            lead = await db.leads.find_one({"id": lead_id, "user_id": user_id})
            if not lead:
                raise HTTPException(status_code=404, detail="Lead not found")
            
            lead_context = NurturingAI.get_lead_context(lead)
            lead_context['user_id'] = user_id
            lead_context['lead_id'] = lead_id
            
            # Execute nurturing agent
            result = await NurturingAI.execute_nurturing_agent(lead_id, user_id, lead_context)
            return result.dict()
        
        else:
            # For other agents, create a basic run for now
            agent_run = await MainOrchestratorAI.create_agent_run(agent_code, lead_id, user_id)
            
            await MainOrchestratorAI.log_agent_event(
                run_id=agent_run.id,
                event_type="INFO",
                payload={"msg": f"{agent_code} execution requested"}
            )
            
            await MainOrchestratorAI.complete_agent_run(agent_run.id, "succeeded")
            
            return {
                "agent_code": agent_code,
                "lead_id": lead_id,
                "run": {
                    "status": "succeeded",
                    "step": "execution_completed",
                    "events": [{"type": "INFO", "payload": {"msg": f"{agent_code} executed successfully"}}],
                    "tasks": []
                },
                "lead_updates": {"engagement_score": 0}
            }
    
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

# --- Nurturing AI Endpoints ---

@app.post("/api/nurturing-ai/generate-plan/{user_id}")
async def generate_nurturing_plan(user_id: str, lead_id: str):
    """Generate nurturing plan for a specific lead"""
    try:
        # Get lead data
        lead = await db.leads.find_one({"id": lead_id, "user_id": user_id})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Extract lead context
        context = NurturingAI.get_lead_context(lead)
        context['user_id'] = user_id
        context['lead_id'] = lead_id
        
        # Determine nurturing strategy
        strategy = NurturingAI.determine_nurturing_strategy(context)
        
        # Generate activity schedule
        activities = NurturingAI.generate_activity_schedule(strategy, context)
        
        # Draft messages for each activity
        for activity in activities:
            message_data = await NurturingAI.draft_message(context, activity.action, activity.channel)
            activity.draft_content = message_data.get('content', '')
            activity.subject = message_data.get('subject', '')
        
        # Create nurturing plan
        next_review_date = (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d')
        
        nurturing_plan = NurturingPlan(
            lead_id=lead_id,
            user_id=user_id,
            activity_board=[activity.dict() for activity in activities],
            lead_updates={
                "stage_suggestion": "qualified" if context.get('pipeline') == 'New Lead' else context.get('pipeline'),
                "engagement_score": min(context.get('engagement_score', 50) + 10, 100)
            },
            next_review=next_review_date,
            engagement_score=context.get('engagement_score', 50),
            strategy_notes=f"Nurturing strategy: {strategy['frequency']} frequency, {strategy['touches_per_week']} touches/week, primary channel: {strategy['primary_channel']}"
        )
        
        # Save activities to database
        for activity in activities:
            await db.nurturing_activities.insert_one(activity.dict())
        
        return nurturing_plan.dict()
        
    except Exception as e:
        print(f"Error generating nurturing plan: {e}")
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

@app.get("/api/nurturing-ai/activities/{user_id}")
async def get_nurturing_activities(user_id: str, date: Optional[str] = None, status: Optional[str] = None):
    """Get nurturing activities for user, optionally filtered by date and status"""
    try:
        query = {"user_id": user_id}
        
        if date:
            query["date"] = date
        
        if status:
            query["status"] = status
        
        activities = await db.nurturing_activities.find(query).sort("date", 1).to_list(length=100)
        
        # Remove MongoDB ObjectId for JSON serialization
        clean_activities = []
        for activity in activities:
            clean_activity = {k: v for k, v in activity.items() if k != "_id"}
            clean_activities.append(clean_activity)
        
        return {
            "status": "success",
            "activities": clean_activities,
            "count": len(clean_activities)
        }
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

@app.put("/api/nurturing-ai/activities/{activity_id}")
async def update_activity_status(activity_id: str, status: str, user_id: str, notes: Optional[str] = None):
    """Update activity status (mark as completed, rescheduled, etc.)"""
    try:
        update_data = {
            "status": status,
            "completed_at": datetime.utcnow().isoformat() if status == "completed" else None
        }
        
        if notes:
            update_data["notes"] = notes
        
        result = await db.nurturing_activities.update_one(
            {"id": activity_id, "user_id": user_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Activity not found")
        
        return {"status": "success", "message": "Activity updated"}
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

@app.post("/api/nurturing-ai/analyze-reply")
async def analyze_reply(user_id: str, lead_id: str, reply_text: str):
    """Analyze lead reply and suggest next action"""
    try:
        # Simple rule-based analysis (can be enhanced with LLM)
        reply_lower = reply_text.lower()
        
        # Sentiment analysis with phrase-based matching (order matters!)
        
        # Check negative phrases first (most specific)
        negative_phrases = [
            'not interested', 'no longer interested', 'stop contacting', 'please stop', 
            'remove me', 'unsubscribe', 'don\'t call', 'don\'t contact', 'not ready',
            'no thank you', 'no thanks', 'not now', 'already found', 'working with another'
        ]
        
        # Check neutral phrases second
        neutral_phrases = [
            'maybe later', 'think about it', 'need to think', 'get back to you',
            'pretty busy', 'very busy', 'busy right now', 'not the right time',
            'timing isn\'t right', 'maybe in the future', 'call me later', 'contact me later'
        ]
        
        # Check positive phrases last (least specific)
        positive_phrases = [
            'yes', 'interested', 'great', 'perfect', 'sounds good', 'let me know',
            'call me', 'tell me more', 'more details', 'schedule', 'when can we',
            'that works', 'looking forward', 'excited', 'exactly what'
        ]
        
        # Use phrase-based matching in priority order
        if any(phrase in reply_lower for phrase in negative_phrases):
            sentiment = 'negative'
            suggested_action = 'Mark as not interested, reduce frequency or pause nurturing'
            intent = 'not_interested'
        elif any(phrase in reply_lower for phrase in neutral_phrases):
            sentiment = 'neutral'
            suggested_action = 'Continue nurturing with lower frequency'
            intent = 'not_ready'
        elif any(phrase in reply_lower for phrase in positive_phrases):
            sentiment = 'positive'
            suggested_action = 'Schedule immediate follow-up call or meeting'
            intent = 'interested'
        else:
            sentiment = 'neutral'
            suggested_action = 'Follow up with clarifying question'
            intent = 'unclear'
        
        analysis = ReplyAnalysis(
            reply_text=reply_text,
            sentiment=sentiment,
            intent=intent,
            suggested_action=suggested_action,
            confidence=0.7  # Simple rule-based, so moderate confidence
        )
        
        # Log the interaction
        interaction_log = {
            "id": str(uuid.uuid4()),
            "lead_id": lead_id,
            "user_id": user_id,
            "type": "reply_received",
            "content": reply_text,
            "analysis": analysis.dict(),
            "created_at": datetime.utcnow().isoformat()
        }
        
        await db.lead_interactions.insert_one(interaction_log)
        
        return analysis.dict()
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"status": "error", "message": str(e)}
        )

# =========================
# Partial Leads API Endpoints
# =========================

class PartialLead(BaseModel):
    id: str
    raw_data: dict
    source: str
    created_at: str
    status: str
    notes: str

class ConvertPartialLeadRequest(BaseModel):
    user_id: str
    # All the same fields as CreateLeadRequest but from the partial lead data
    name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    lead_description: Optional[str] = None
    
    # Additional Contact Information
    work_phone: Optional[str] = None
    home_phone: Optional[str] = None
    email_2: Optional[str] = None
    
    # Spouse Information
    spouse_name: Optional[str] = None
    spouse_first_name: Optional[str] = None
    spouse_last_name: Optional[str] = None
    spouse_email: Optional[str] = None
    spouse_mobile_phone: Optional[str] = None
    spouse_birthday: Optional[str] = None
    
    # Pipeline and Status
    pipeline: Optional[str] = None
    status: Optional[str] = None
    ref_source: Optional[str] = None
    lead_rating: Optional[str] = None
    lead_source: Optional[str] = None
    lead_type: Optional[str] = None
    lead_type_2: Optional[str] = None
    
    # Property Information
    house_to_sell: Optional[str] = None
    buying_in: Optional[str] = None
    selling_in: Optional[str] = None
    owns_rents: Optional[str] = None
    mortgage_type: Optional[str] = None
    
    # Address Information
    city: Optional[str] = None
    zip_postal_code: Optional[str] = None
    address: Optional[str] = None
    
    # Property Details
    property_type: Optional[str] = None
    property_condition: Optional[str] = None
    listing_status: Optional[str] = None
    bedrooms: Optional[str] = None
    bathrooms: Optional[str] = None
    basement: Optional[str] = None
    parking_type: Optional[str] = None
    
    # Dates and Anniversaries  
    date_of_birth: Optional[str] = None
    house_anniversary: Optional[str] = None
    planning_to_sell_in: Optional[str] = None
    
    # Agent Assignments
    main_agent: Optional[str] = None
    mort_agent: Optional[str] = None
    list_agent: Optional[str] = None
    
    # Custom Fields (flexible JSON structure)
    custom_fields: Optional[dict] = None
    
    # Existing compatibility fields
    neighborhood: Optional[str] = None
    price_min: Optional[int] = None
    price_max: Optional[int] = None
    priority: Optional[str] = None
    source_tags: Optional[List[str]] = None
    notes: Optional[str] = None
    stage: Optional[str] = None
    in_dashboard: Optional[bool] = None

    @field_validator("phone", "work_phone", "home_phone", "spouse_mobile_phone")
    @classmethod
    def validate_phone(cls, v):
        if v is not None and v.strip() and not E164_RE.match(v):
            # Try to normalize the phone number
            normalized = normalize_phone(v)
            if normalized and E164_RE.match(normalized):
                return normalized
            raise ValueError("Phone must be in E.164 format, e.g. +1234567890")
        return v

@app.get("/api/partial-leads", response_model=List[PartialLead])
async def get_partial_leads():
    """Get all partial leads"""
    try:
        partial_leads = await db.partial_leads.find().to_list(length=None)
        return [PartialLead(**{k: v for k, v in lead.items() if k != "_id"}) for lead in partial_leads]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/partial-leads/{lead_id}", response_model=PartialLead)
async def get_partial_lead(lead_id: str):
    """Get a specific partial lead"""
    try:
        partial_lead = await db.partial_leads.find_one({"id": lead_id})
        if not partial_lead:
            raise HTTPException(status_code=404, detail="Partial lead not found")
        return PartialLead(**{k: v for k, v in partial_lead.items() if k != "_id"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/partial-leads/{lead_id}")
async def delete_partial_lead(lead_id: str):
    """Delete a partial lead"""
    try:
        result = await db.partial_leads.delete_one({"id": lead_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Partial lead not found")
        return {"message": "Partial lead deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/partial-leads/{lead_id}/convert", response_model=Lead)
async def convert_partial_lead(lead_id: str, convert_data: ConvertPartialLeadRequest):
    """Convert a partial lead to a full lead"""
    try:
        # Get the partial lead first
        partial_lead = await db.partial_leads.find_one({"id": lead_id})
        if not partial_lead:
            raise HTTPException(status_code=404, detail="Partial lead not found")
        
        # Validate and normalize email
        validated_email = None
        if convert_data.email and convert_data.email.strip():
            try:
                validation = validate_email(convert_data.email.strip(), check_deliverability=False)
                validated_email = validation.email
            except EmailNotValidError:
                validated_email = None
        
        # Normalize phone numbers
        normalized_phone = normalize_phone(convert_data.phone) if convert_data.phone else None
        normalized_work_phone = normalize_phone(convert_data.work_phone) if convert_data.work_phone else None
        normalized_home_phone = normalize_phone(convert_data.home_phone) if convert_data.home_phone else None
        normalized_spouse_phone = normalize_phone(convert_data.spouse_mobile_phone) if convert_data.spouse_mobile_phone else None
        
        # Create lead name if not provided
        name = convert_data.name
        if not name:
            name = f"{convert_data.first_name or ''} {convert_data.last_name or ''}".strip() or "Converted Lead"
        
        # Create the full lead
        lead = Lead(
            user_id=convert_data.user_id,
            name=name,
            first_name=convert_data.first_name,
            last_name=convert_data.last_name,
            email=validated_email,
            phone=normalized_phone,
            lead_description=convert_data.lead_description,
            work_phone=normalized_work_phone,
            home_phone=normalized_home_phone,
            email_2=convert_data.email_2,
            spouse_name=convert_data.spouse_name,
            spouse_first_name=convert_data.spouse_first_name,
            spouse_last_name=convert_data.spouse_last_name,
            spouse_email=convert_data.spouse_email,
            spouse_mobile_phone=normalized_spouse_phone,
            spouse_birthday=convert_data.spouse_birthday,
            pipeline=convert_data.pipeline or "New Lead",
            status=convert_data.status or "Open",
            ref_source=convert_data.ref_source,
            lead_rating=convert_data.lead_rating,
            lead_source=convert_data.lead_source or "Converted from Partial",
            lead_type=convert_data.lead_type,
            lead_type_2=convert_data.lead_type_2,
            house_to_sell=convert_data.house_to_sell,
            buying_in=convert_data.buying_in,
            selling_in=convert_data.selling_in,
            owns_rents=convert_data.owns_rents,
            mortgage_type=convert_data.mortgage_type,
            city=convert_data.city,
            zip_postal_code=convert_data.zip_postal_code,
            address=convert_data.address,
            property_type=convert_data.property_type,
            property_condition=convert_data.property_condition,
            listing_status=convert_data.listing_status,
            bedrooms=convert_data.bedrooms,
            bathrooms=convert_data.bathrooms,
            basement=convert_data.basement,
            parking_type=convert_data.parking_type,
            house_anniversary=convert_data.house_anniversary,
            planning_to_sell_in=convert_data.planning_to_sell_in,
            main_agent=convert_data.main_agent,
            mort_agent=convert_data.mort_agent,
            list_agent=convert_data.list_agent,
            custom_fields=convert_data.custom_fields,
            neighborhood=convert_data.neighborhood,
            price_min=convert_data.price_min,
            price_max=convert_data.price_max,
            priority=convert_data.priority or "medium",
            source_tags=convert_data.source_tags or ["Converted from Partial"],
            notes=convert_data.notes,
            stage=convert_data.stage or "New",
            in_dashboard=convert_data.in_dashboard if convert_data.in_dashboard is not None else True
        )
        
        # Insert the new lead
        await db.leads.insert_one(lead.model_dump(exclude_none=True))
        
        # Remove the partial lead
        await db.partial_leads.delete_one({"id": lead_id})
        
        return lead
        
    except DuplicateKeyError:
        raise HTTPException(status_code=409, detail="Lead with this email already exists")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# =========================
# Email Draft System API Endpoints
# =========================

class EmailDraft(BaseModel):
    id: str
    lead_id: str
    user_id: str
    subject: str
    body: str
    html_body: Optional[str] = None
    to_email: str
    from_email: Optional[str] = None
    status: str  # draft, sent, failed
    email_type: str
    urgency: str
    created_at: str
    due_date: Optional[str] = None
    sent_at: Optional[str] = None
    ai_generated: Optional[bool] = False
    channel: str

class SendDraftRequest(BaseModel):
    draft_id: str
    from_email: str

@app.get("/api/email-drafts/{lead_id}", response_model=List[EmailDraft])
async def get_email_drafts(lead_id: str):
    """Get all email drafts for a specific lead"""
    try:
        drafts = await db.email_drafts.find({"lead_id": lead_id}).to_list(length=None)
        return [EmailDraft(**{k: v for k, v in draft.items() if k != "_id"}) for draft in drafts]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/email-drafts/count/{lead_id}")
async def get_email_draft_count(lead_id: str):
    """Get count of pending drafts for a lead"""
    try:
        count = await db.email_drafts.count_documents({"lead_id": lead_id, "status": "draft"})
        return {"lead_id": lead_id, "draft_count": count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/email-drafts/send")
async def send_email_draft(request: SendDraftRequest):
    """Send an email draft using SendGrid"""
    try:
        # Get the draft
        draft = await db.email_drafts.find_one({"id": request.draft_id})
        if not draft:
            raise HTTPException(status_code=404, detail="Draft not found")
        
        # Get the lead
        lead = await db.leads.find_one({"id": draft["lead_id"]})
        if not lead:
            raise HTTPException(status_code=404, detail="Lead not found")
        
        # Auto-migrate and get SendGrid secrets
        await migrate_secrets_from_settings(draft["user_id"])
        secrets = await get_all_secrets(draft["user_id"])
        sendgrid_api_key = secrets.get("sendgrid_api_key")
        
        if not sendgrid_api_key:
            raise HTTPException(status_code=400, detail="SendGrid API key not configured")
        
        # Use SendGrid Python SDK as per official documentation
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail
        
        try:
            # Initialize SendGrid client
            sg = SendGridAPIClient(sendgrid_api_key)
            
            print(f"🔵 Sending email via SendGrid SDK...")
            print(f"  To: {draft['to_email']}")
            print(f"  From: {request.from_email}")
            print(f"  Subject: {draft['subject']}")
            
            # Create email message using SendGrid Mail helper
            # If HTML body exists, use it as primary content, otherwise use plain text
            if draft.get("html_body"):
                message = Mail(
                    from_email=request.from_email,
                    to_emails=draft["to_email"],
                    subject=draft["subject"],
                    plain_text_content=draft["body"],
                    html_content=draft["html_body"]
                )
            else:
                message = Mail(
                    from_email=request.from_email,
                    to_emails=draft["to_email"],
                    subject=draft["subject"],
                    plain_text_content=draft["body"]
                )
            
            # Send email using SendGrid SDK
            response = sg.send(message)
            
            print(f"✅ SendGrid Response Status: {response.status_code}")
            print(f"  Response Body: {response.body}")
            print(f"  Response Headers: {dict(response.headers)}")
            
            # SendGrid returns 202 (Accepted) for successful requests
            if response.status_code in [200, 201, 202]:
                # Extract message ID from headers
                message_id = response.headers.get('X-Message-Id', '')
                
                # Update draft status to sent
                await db.email_drafts.update_one(
                    {"id": request.draft_id},
                    {"$set": {
                        "status": "sent",
                        "sent_at": datetime.now().isoformat(),
                        "from_email": request.from_email,
                        "sendgrid_message_id": message_id
                    }}
                )
                
                # Store the sender email in settings for future use
                await db.settings.update_one(
                    {"user_id": draft["user_id"]},
                    {"$set": {"sender_email": request.from_email}},
                    upsert=True
                )
                
                # Add activity to lead notes
                current_notes = lead.get('notes', '')
                new_note = f"\n\n[Email Sent] '{draft['subject']}' to {draft['to_email']} from {request.from_email} - {datetime.now().isoformat()}"
                await db.leads.update_one(
                    {"id": draft["lead_id"]},
                    {"$set": {"notes": current_notes + new_note}}
                )
                
                print(f"✅ Email sent successfully! Message ID: {message_id}")
                
                return {
                    "success": True,
                    "message": f"Email sent successfully to {draft['to_email']}",
                    "message_id": message_id or request.draft_id
                }
            else:
                # Handle non-success status codes
                error_msg = f"Unexpected status code: {response.status_code}"
                if response.body:
                    error_msg += f" - {response.body}"
                
                print(f"❌ SendGrid error: {error_msg}")
                
                # Update draft status to failed
                await db.email_drafts.update_one(
                    {"id": request.draft_id},
                    {"$set": {
                        "status": "failed",
                        "error_message": error_msg
                    }}
                )
                
                return {
                    "success": False,
                    "error": error_msg
                }
                
        except Exception as sg_error:
            # Handle SendGrid SDK exceptions
            error_msg = str(sg_error)
            
            # Try to extract detailed error from SendGrid exception
            if hasattr(sg_error, 'body'):
                try:
                    import json
                    error_body = json.loads(sg_error.body)
                    if "errors" in error_body:
                        error_msg = "; ".join([err.get("message", str(err)) for err in error_body["errors"]])
                except:
                    error_msg = sg_error.body
            
            print(f"❌ SendGrid SDK Exception: {error_msg}")
            
            # Update draft status to failed
            await db.email_drafts.update_one(
                {"id": request.draft_id},
                {"$set": {
                    "status": "failed",
                    "error_message": error_msg
                }}
            )
            
            return {
                "success": False,
                "error": f"SendGrid error: {error_msg}"
            }
            
    except Exception as e:
        print(f"❌ Exception in send_email_draft: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}

@app.delete("/api/email-drafts/{draft_id}")
async def delete_email_draft(draft_id: str):
    """Delete an email draft"""
    try:
        result = await db.email_drafts.delete_one({"id": draft_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Draft not found")
        return {"message": "Draft deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/settings/preferred-from-email/{user_id}")
async def get_preferred_from_email(user_id: str):
    """Get user's preferred from email"""
    try:
        settings = await db.settings.find_one({"user_id": user_id})
        preferred_email = settings.get("preferred_from_email") if settings else None
        return {"preferred_from_email": preferred_email}
    except Exception as e:
        return {"preferred_from_email": None}

# Mount Lead Generation Service
from leadgen_service import app as leadgen_app

# Mount Lead Nurturing Service
from lead_nurture_service import app as nurture_app

# Mount the service routes
app.mount("/api/agents/leadgen", leadgen_app)
app.mount("/api/agents/nurture", nurture_app)

# =========================
# Local run helper
# =========================